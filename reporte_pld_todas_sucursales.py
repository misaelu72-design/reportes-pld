"""
Reporte PLD — Todas las Sucursales (11 hojas)
=============================================
Genera el archivo Reporte_PLD_TODAS_SUCURSALES_<fecha>.xlsx

Uso:
    1. Edita la sección CONFIGURACIÓN con tus rutas y fechas
    2. Ejecuta: python reporte_pld_todas_sucursales.py
"""

import re, math, json
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ══════════════════════════════════════════════════════════════
# CONFIGURACIÓN — edita solo esta sección en cada nuevo corte
# ══════════════════════════════════════════════════════════════

# ── Carpetas base — no cambiar ──────────────────────────────
BASE = Path(r"C:\Users\Misael Urbina\OneDrive - CECOM\Documentos\PORTAL WEB PLD")

CARPETA_ANTERIOR = BASE / "ANTERIOR"   # archivos de la semana pasada (se copian automático)
CARPETA_ABRIL    = BASE / "04-26"        # 26 de Abril — punto de referencia fijo (nunca cambia)
CARPETA_ACTUAL   = BASE / "ACTUAL"     # archivos de esta semana (tú los pones aquí)
CARPETA_SALIDA   = BASE / "COMPARATIVAS"  # reportes generados

# Etiquetas de fecha para los encabezados
FECHA_ANTERIOR_CORTA = "17-May"
FECHA_ACTUAL_CORTA   = "24-May"
FECHA_ANTERIOR_LARGA = "17 de Mayo 2026"
FECHA_ACTUAL_LARGA   = "24 de Mayo 2026"

# Rangos de semana para H8/H9 (semana anterior y semana nueva)
SEMANA_ANT_INICIO = pd.Timestamp("2026-05-11")
SEMANA_ANT_FIN    = pd.Timestamp("2026-05-17")
SEMANA_NVA_INICIO = pd.Timestamp("2026-05-18")
SEMANA_NVA_FIN    = pd.Timestamp("2026-05-24")

# YTD — clientes creados desde esta fecha
YTD_INICIO = pd.Timestamp("2026-01-01")

# Rango del corte para el reporte de operaciones
# Se calcula automáticamente desde las carpetas anterior y actual
FECHA_CORTE_INICIO = SEMANA_ANT_INICIO   # inicio de la semana anterior
FECHA_CORTE_FIN    = SEMANA_NVA_FIN      # fin de la semana nueva

# ── Correo de notificación ──────────────────────────────────
CORREO_REMITENTE     = "misael.urbina@mycecom.com"
CORREO_DESTINATARIOS = [
    "misael.urbina@mycecom.com",
    "nydia.salas@mycecom.com",
    "perla.rojas@mycecom.com",
]
# Contraseña Outlook — si se deja vacía el script la pedirá al correr
CORREO_PASSWORD = ""

# ── WhatsApp (CallMeBot) ─────────────────────────────────────
# Pasos para activar:
# 1. Guarda +34 644 65 21 69 en tus contactos
# 2. Mándale por WhatsApp: "I allow callmebot to send me messages"
# 3. Te responde con tu API key — ponla aquí
WHATSAPP_APIKEY  = ""        # ← pega aquí tu API key de CallMeBot
WHATSAPP_NUMEROS = [
    "+5281XXXXXXXX",          # ← tu número con código de país, sin espacios
    # Agrega más números aquí:
    # "+5281YYYYYYYY",
]
URL_PORTAL = "https://misaelu72-design.github.io/reportes-pld/"

# Sucursales a procesar (el script busca el archivo automáticamente por nombre)
SUCURSALES = [
    "Ajijic", "Allende", "Esmeralda", "Guadalupe",
    "Irapuato", "Mitras", "Neza", "Satelite", "Sendero",
]

# ══════════════════════════════════════════════════════════════
# CONSTANTES (no editar)
# ══════════════════════════════════════════════════════════════

FECHA_ID_VENCIDA = pd.Timestamp("2025-12-31")
HOJA_DATOS       = "KYCTotal"

OPS_STATUS = [
    "Bloqueado operaciones", "Info actualizada operaciones",
    "Registro completo", "Verificado operaciones",
]
KYC_STATUS = [
    "Bloqueado PLD", "Info actualizada PLD",
    "Por validar PLD", "Verificado PLD",
]
TODOS_STATUS = OPS_STATUS + KYC_STATUS + ["Registro express"]

CATEGORIAS_NOTAS = [
    "CLIENTE EXCLUIDO SEGOB", "CLIENTE VETADO", "CLIENTE VERIFICADO",
    "ID VENCIDA", "ID NO VALIDA", "ID INCOMPLETA VOLVER A ESCANEAR",
    "FALTA COMPROBANTE DE DOMICILIO", "FALTA OCUPACION", "DATOS DE CONTACTO",
    "SIN HUELLAS DIGITALES", "SIN AVISO FIRMADO", "SIN AVISO FIRMADO - SISTEMA",
    "FALTA NACIONALIDAD", "CODIGO POSTAL DISTINTO", "ERROR DE RFC",
    "OTROS", "SIN NOTA",
]

C = {
    "azul_marino":  "1F3864", "azul_medio":   "2E75B6",
    "verde_oscuro": "375623", "naranja":       "E26B0A",
    "azul_claro":   "BDD7EE", "verde_claro":  "E2EFDA",
    "verde_varia":  "C6EFCE", "rojo_varia":   "FFC7CE",
    "salmon":       "FCE4D6", "rosa":          "F4CCCC",
    "amarillo":     "FFF2CC", "gris_claro":   "F2F2F2",
    "blanco":       "FFFFFF",
}
COLOR_NIVEL = {
    "Corazon": "DEEBF7", "Corazon Plus": "FFF2CC",
    "Diamante": "E2EFDA", "Diamante Plus": "FCE4D6",
    "Espada": "EAF0FB",  "Espada Plus": "E2EFDA",
    "Foliatti Fan": "FFFFFF", "Trebol": "D9EAD3", "SIN NIVEL": "F2F2F2",
}
COLOR_STATUS = {
    "Bloqueado PLD": "F4CCCC",
    "Bloqueado operaciones": "FCE4D6",
    "Info actualizada PLD": "E2EFDA",
    "Info actualizada operaciones": "BDD7EE",
    "Por validar PLD": "DEEBF7",
    "Registro completo": "EAF0FB",
    "Registro express": "FFF2CC",
    "Verificado PLD": "E2EFDA",
    "Verificado operaciones": "EAF0FB",
}

# ══════════════════════════════════════════════════════════════
# HELPERS DE ESTILO
# ══════════════════════════════════════════════════════════════

def fill(h):   return PatternFill("solid", fgColor=h)
def font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def s(cell, bg=None, bold=False, color="000000", size=10,
      h="left", v="center", wrap=False):
    if bg: cell.fill = fill(bg)
    cell.font      = font(bold=bold, color=color, size=size)
    cell.alignment = aln(h=h, v=v, wrap=wrap)

def titulo(ws, txt, fila, ncols, size=14):
    ws.merge_cells(start_row=fila, start_column=1,
                   end_row=fila,   end_column=ncols)
    c = ws.cell(fila, 1, txt)
    s(c, bg=C["azul_marino"], bold=True, color="FFFFFF", size=size)
    ws.row_dimensions[fila].height = 36

def seccion(ws, txt, fila, ncols, bg="azul_marino", size=11):
    ws.merge_cells(start_row=fila, start_column=1,
                   end_row=fila,   end_column=ncols)
    c = ws.cell(fila, 1, txt)
    s(c, bg=C[bg], bold=True, color="FFFFFF", size=size)
    ws.row_dimensions[fila].height = 22

def enc(ws, headers, fila, bg="azul_marino"):
    for j, h in enumerate(headers, 1):
        c = ws.cell(fila, j, h)
        s(c, bg=C[bg], bold=True, color="FFFFFF", size=10,
          h="center" if j > 1 else "left", wrap=True)
    ws.row_dimensions[fila].height = 28

def var_cell(cell, val):
    v = str(val)
    if v.startswith("+") and v != "+":
        cell.fill = fill(C["verde_varia"])
        cell.font = Font(bold=True, color="375623", size=10, name="Arial")
    elif v.startswith("-") and v not in ["-", "0"]:
        cell.fill = fill(C["rojo_varia"])
        cell.font = Font(bold=True, color="9C0006", size=10, name="Arial")
    cell.alignment = aln(h="center")

def var_str(n):
    if n == 0: return "0"
    return f"+{n:,}" if n > 0 else f"{n:,}"

def pct(n, total):
    return f"{n/total*100:.1f}%" if total else "0.0%"

# ══════════════════════════════════════════════════════════════
# LECTURA Y NORMALIZACIÓN
# ══════════════════════════════════════════════════════════════

STATUS_MAPA = {
    "registro express": "Registro express",
    "bloqueado por operaciones": "Bloqueado operaciones",
    "bloqueo por operaciones": "Bloqueado operaciones",
    "información actualizada operaciones": "Info actualizada operaciones",
    "informacion actualizada operaciones": "Info actualizada operaciones",
    "registro completo": "Registro completo",
    "verificado por operaciones": "Verificado operaciones",
    "bloqueado por kyc": "Bloqueado PLD",
    "bloqueo por kyc": "Bloqueado PLD",
    "información actualizada kyc": "Info actualizada PLD",
    "informacion actualizada kyc": "Info actualizada PLD",
    "pendente validacion kyc": "Por validar PLD",
    "pendiente validacion kyc": "Por validar PLD",
    "pendiente validación kyc": "Por validar PLD",
    "verificado por kyc": "Verificado PLD",
}

def limpiar(x):
    if x is None or (isinstance(x, float) and math.isnan(x)): return ""
    return str(x).strip()

def norm_status(x):
    t = re.sub(r"\s+", " ", limpiar(x).lower())
    return STATUS_MAPA.get(t, limpiar(x))

def clasificar_nota(nota):
    t = re.sub(r"\s+", " ", limpiar(nota).upper())
    if not t:                                             return "SIN NOTA"
    if "SEGOB" in t or "EXCLUIDO" in t:                  return "CLIENTE EXCLUIDO SEGOB"
    if "VETADO" in t:                                     return "CLIENTE VETADO"
    if "VERIFICADO" in t:                                 return "CLIENTE VERIFICADO"
    if "ID VENCIDA" in t or "INE VENCIDA" in t:           return "ID VENCIDA"
    if "ID NO VALIDA" in t or "INE NO VALIDA" in t:       return "ID NO VALIDA"
    if "ID INCOMPLETA" in t or "VOLVER A ESCANEAR" in t:  return "ID INCOMPLETA VOLVER A ESCANEAR"
    if "COMPROBANTE" in t or "DOMICILIO" in t:            return "FALTA COMPROBANTE DE DOMICILIO"
    if "OCUPACION" in t or "OCUPACIÓN" in t:              return "FALTA OCUPACION"
    if "CONTACTO" in t or "CORREO" in t or "TELEFONO" in t or "TELÉFONO" in t:
        return "DATOS DE CONTACTO"
    if "HUELLA" in t:                                     return "SIN HUELLAS DIGITALES"
    if "AVISO" in t and "SISTEMA" in t:                   return "SIN AVISO FIRMADO - SISTEMA"
    if "AVISO" in t:                                      return "SIN AVISO FIRMADO"
    if "NACIONALIDAD" in t:                               return "FALTA NACIONALIDAD"
    if "CODIGO POSTAL" in t or "CÓDIGO POSTAL" in t or " CP " in f" {t} ":
        return "CODIGO POSTAL DISTINTO"
    if "RFC" in t:                                        return "ERROR DE RFC"
    return "OTROS"

def es_si(x):
    return limpiar(x).upper() in ["SI", "SÍ", "1", "1.0", "TRUE", "VERDADERO", "YES"]

def buscar_col(df, opts):
    mapa = {str(c).strip().upper().replace("\n", " ").replace("_", " "): c
            for c in df.columns}
    for o in opts:
        r = mapa.get(o.strip().upper().replace("_", " "))
        if r: return r
    return None

def buscar_archivo(carpeta: Path, sucursal: str) -> Path | None:
    """Busca el xlsx de una sucursal en la carpeta sin importar el nombre exacto."""
    nombre_norm = sucursal.upper().replace(" ", "")
    for archivo in sorted(carpeta.glob("*.xlsx")):
        if archivo.name.startswith("~$"):
            continue
        stem_norm = archivo.stem.upper().replace(" ", "").replace("_", "")
        if nombre_norm in stem_norm:
            return archivo
    return None

def leer_df(path: Path, sucursal: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=HOJA_DATOS)
    df.columns = [str(c).strip() for c in df.columns]

    col_id      = buscar_col(df, ["PLAYER_ID", "ID_JUGADOR", "Player ID", "ID"])
    col_status  = buscar_col(df, ["STATUS", "ESTATUS", "ESTADO"])
    col_notas   = buscar_col(df, ["NOTAS", "NOTA", "OBSERVACIONES"])
    col_exp     = buscar_col(df, ["EXPIRATION_IFE", "EXPIRATION IFE", "VENCIMIENTO_IFE"])
    col_aviso   = buscar_col(df, ["AVISO_FIRMADO", "AVISO FIRMADO", "AVISO_PRIVACIDAD"])
    col_huella  = buscar_col(df, ["HUELLA", "HUELLA DIGITAL", "HUELLA_DIGITAL"])
    col_nivel   = buscar_col(df, ["NIVEL_LEALTAD", "NIVEL LEALTAD", "PLAYER_LEVEL_NAME"])
    col_nombre  = buscar_col(df, ["NOMBRE_JUGADOR", "NOMBRE", "NAME"])
    col_created = buscar_col(df, ["DATE_CREATED_NEW_RECORD", "DATE CREATED NEW RECORD",
                                  "FECHA_CREACION", "FECHA CREACION"])
    col_updated = buscar_col(df, ["DATE_CREATED_UPDATE_RECORD", "DATE CREATED UPDATE RECORD",
                                  "FECHA_ACTUALIZACION"])

    if not col_id:     raise ValueError(f"Sin PLAYER_ID en {path.name}")
    if not col_status: raise ValueError(f"Sin STATUS en {path.name}")

    df["_SUCURSAL"] = sucursal
    df["_ID"]       = df[col_id].astype(str).str.strip()
    df["_STATUS"]   = df[col_status].apply(norm_status)
    df["_NOTAS"]    = df[col_notas].apply(limpiar) if col_notas \
                      else pd.Series("", index=df.index)
    df["_NOTA_CAT"] = df["_NOTAS"].apply(clasificar_nota)
    df["_NIVEL"]    = df[col_nivel].apply(limpiar).replace("", "SIN NIVEL") if col_nivel \
                      else pd.Series("SIN NIVEL", index=df.index)
    df["_NOMBRE"]   = df[col_nombre].apply(limpiar) if col_nombre \
                      else pd.Series("", index=df.index)
    df["_EXP"]      = pd.to_datetime(df[col_exp], errors="coerce") if col_exp else pd.NaT
    df["_ID_VENC"]  = df["_EXP"].isna() | (df["_EXP"] <= FECHA_ID_VENCIDA)
    df["_AVISO"]    = df[col_aviso].apply(es_si) if col_aviso \
                      else pd.Series(False, index=df.index)
    df["_HUELLA"]   = df[col_huella].apply(es_si) if col_huella \
                      else pd.Series(False, index=df.index)
    df["_CREATED"]  = pd.to_datetime(df[col_created], errors="coerce") if col_created \
                      else pd.NaT
    df["_UPDATED"]  = pd.to_datetime(df[col_updated], errors="coerce") if col_updated \
                      else pd.NaT

    return df[df["_ID"].str.len() > 0].drop_duplicates(subset=["_ID"], keep="last")

def cnt(df, st):
    return int((df["_STATUS"] == st).sum())

# ══════════════════════════════════════════════════════════════
# BLOQUE DE ESTATUS REUTILIZABLE (usado en H4)
# ══════════════════════════════════════════════════════════════

def bloque_estatus(ws, fila_ini, lbl_total, df_ant, df_act):
    enc(ws, ["ESTATUS / GRUPO", FECHA_ANTERIOR_CORTA,
             FECHA_ACTUAL_CORTA, "VARIACIÓN"], fila_ini)
    ws.row_dimensions[fila_ini].height = 24

    ops_a = sum(cnt(df_ant, st) for st in OPS_STATUS)
    ops_c = sum(cnt(df_act, st) for st in OPS_STATUS)
    kyc_a = sum(cnt(df_ant, st) for st in KYC_STATUS)
    kyc_c = sum(cnt(df_act, st) for st in KYC_STATUS)

    filas = [
        (lbl_total, len(df_ant), len(df_act), C["naranja"], True, "FFFFFF"),
        None,
        ("REGISTRO EXPRESS",
         cnt(df_ant,"Registro express"), cnt(df_act,"Registro express"),
         C["azul_medio"], True, "FFFFFF"),
        None,
        ("GRUPO: OPERACIONES", ops_a, ops_c, C["azul_medio"], True, "FFFFFF"),
        ("      Bloqueado Operaciones",
         cnt(df_ant,"Bloqueado operaciones"),
         cnt(df_act,"Bloqueado operaciones"),
         C["azul_claro"], False, "000000"),
        ("      Info Act Operaciones",
         cnt(df_ant,"Info actualizada operaciones"),
         cnt(df_act,"Info actualizada operaciones"),
         C["azul_claro"], False, "000000"),
        ("      Registro Completo",
         cnt(df_ant,"Registro completo"), cnt(df_act,"Registro completo"),
         C["azul_claro"], False, "000000"),
        ("      Verificado Operaciones",
         cnt(df_ant,"Verificado operaciones"),
         cnt(df_act,"Verificado operaciones"),
         C["azul_claro"], False, "000000"),
        None,
        ("GRUPO: KYC-PLD", kyc_a, kyc_c, C["verde_oscuro"], True, "FFFFFF"),
        ("      Bloqueado PLD",
         cnt(df_ant,"Bloqueado PLD"), cnt(df_act,"Bloqueado PLD"),
         C["verde_claro"], False, "000000"),
        ("      Info Act PLD",
         cnt(df_ant,"Info actualizada PLD"),
         cnt(df_act,"Info actualizada PLD"),
         C["verde_claro"], False, "000000"),
        ("      Por validar PLD",
         cnt(df_ant,"Por validar PLD"),
         cnt(df_act,"Por validar PLD"),
         C["verde_claro"], False, "000000"),
        ("      Verificado PLD",
         cnt(df_ant,"Verificado PLD"), cnt(df_act,"Verificado PLD"),
         C["verde_claro"], False, "000000"),
        None,
        ("TOTAL (OPERACIONES + KYC-PLD)",
         ops_a+kyc_a, ops_c+kyc_c, C["naranja"], True, "FFFFFF"),
    ]

    f = fila_ini + 1
    for item in filas:
        if item is None:
            ws.row_dimensions[f].height = 8; f += 1; continue
        lbl, va, vc, bg, bold, col = item
        vr = var_str(vc - va)
        for j, v in enumerate([lbl, va, vc, vr], 1):
            c = ws.cell(f, j, v)
            if j == 4 and not bold:
                var_cell(c, vr)
            else:
                s(c, bg=bg, bold=bold, color=col, size=10,
                  h="left" if j == 1 else "center")
            if j == 4 and bold:
                s(c, bg=bg, bold=True, color=col, size=10, h="center")
        ws.row_dimensions[f].height = 20
        f += 1
    return f

# ══════════════════════════════════════════════════════════════
# H1 — CAMBIO DE ESTATUS
# ══════════════════════════════════════════════════════════════

def hacer_h1(wb, todos_cambios, todos_nuevos):
    ws = wb.create_sheet("H1 - Cambio de Estatus")
    ncols = 5
    titulo(ws, f"RESUMEN DE CAMBIO DE ESTATUS  |  "
               f"{FECHA_ANTERIOR_CORTA} vs {FECHA_ACTUAL_CORTA}", 1, ncols)

    total_c = len(todos_cambios)
    ws.merge_cells("A2:E2")
    c = ws.cell(2, 1, f"Total cambios: {total_c:,}   |   "
                      f"Clientes nuevos: {len(todos_nuevos):,}")
    s(c, bg=C["azul_medio"], bold=True, color="FFFFFF", size=11)
    ws.row_dimensions[2].height = 20

    # Transiciones globales
    seccion(ws, "▶  TRANSICIONES GLOBALES DE ESTATUS", 4, ncols)
    enc(ws, ["#", f"STATUS ANTERIOR ({FECHA_ANTERIOR_CORTA})",
             f"STATUS NUEVO ({FECHA_ACTUAL_CORTA})",
             "CANTIDAD", "% DEL TOTAL"], 5)

    trans = (todos_cambios
             .groupby(["_STATUS_ANT", "_STATUS_ACT"]).size()
             .reset_index(name="N")
             .sort_values("N", ascending=False)
             .reset_index(drop=True))

    fila = 6
    for idx, row in trans.iterrows():
        bg = COLOR_STATUS.get(row["_STATUS_ANT"], C["blanco"])
        pct_v = f"{row['N']/total_c*100:.1f}%" if total_c else "0.0%"
        for j, v in enumerate([idx+1, row["_STATUS_ANT"], row["_STATUS_ACT"],
                                int(row["N"]), pct_v], 1):
            c = ws.cell(fila, j, v)
            s(c, bg=bg, size=10, h="center" if j in [1, 4, 5] else "left")
        ws.row_dimensions[fila].height = 18
        fila += 1

    ws.cell(fila, 2, "TOTAL CAMBIOS")
    ws.cell(fila, 4, total_c)
    ws.cell(fila, 5, "100.0%")
    for j in range(1, 6):
        s(ws.cell(fila, j), bg=C["azul_marino"], bold=True,
          color="FFFFFF", size=10, h="center" if j != 2 else "left")
    ws.row_dimensions[fila].height = 20
    fila += 2

    # Cambios por sucursal
    seccion(ws, "▶  CAMBIOS POR SUCURSAL", fila, ncols)
    fila += 1
    enc(ws, ["SUCURSAL", f"STATUS ANTERIOR ({FECHA_ANTERIOR_CORTA})",
             f"STATUS NUEVO ({FECHA_ACTUAL_CORTA})", "CANTIDAD", ""], fila)
    fila += 1

    for suc in SUCURSALES:
        df_s = todos_cambios[todos_cambios["_SUCURSAL"] == suc]
        if df_s.empty: continue

        ws.merge_cells(start_row=fila, start_column=1,
                       end_row=fila,   end_column=3)
        c = ws.cell(fila, 1, f"  {suc.upper()}  — {len(df_s):,} cambios")
        ws.cell(fila, 4, len(df_s))
        s(c, bg=C["azul_claro"], bold=True, color=C["azul_marino"], size=10)
        s(ws.cell(fila, 4), bg=C["azul_claro"], bold=True,
          color=C["azul_marino"], size=10, h="center")
        ws.row_dimensions[fila].height = 20
        fila += 1

        trans_suc = (df_s.groupby(["_STATUS_ANT", "_STATUS_ACT"]).size()
                     .reset_index(name="N").sort_values("N", ascending=False))
        for _, row in trans_suc.iterrows():
            bg = COLOR_STATUS.get(row["_STATUS_ANT"], C["blanco"])
            for j, v in enumerate(["", row["_STATUS_ANT"], row["_STATUS_ACT"],
                                    int(row["N"]), ""], 1):
                c = ws.cell(fila, j, v)
                s(c, bg=bg, size=10, h="center" if j == 4 else "left")
            ws.row_dimensions[fila].height = 18
            fila += 1
        fila += 1

    # ── Clientes nuevos (sección al final de H1, igual que el reporte de referencia)
    fila += 1

    # Encabezado con total entre paréntesis
    ws.merge_cells(start_row=fila, start_column=1,
                   end_row=fila,   end_column=ncols)
    c = ws.cell(fila, 1,
                f"▶  CLIENTES NUEVOS EN KYCTOTAL_TODAS_SUCURSALES "
                f"{FECHA_ACTUAL_CORTA.upper()}  ({len(todos_nuevos):,} clientes)")
    s(c, bg=C["azul_marino"], bold=True, color="FFFFFF", size=11)
    ws.row_dimensions[fila].height = 24
    fila += 1

    # Encabezados de columna
    enc(ws, ["SUCURSAL", "STATUS", "CANTIDAD", "", ""], fila)
    ws.row_dimensions[fila].height = 22
    fila += 1

    # Colores alternos por sucursal (igual que en el reporte de referencia)
    COLORES_SUCURSAL = [
        ("2E75B6", "FFFFFF"),  # azul medio
        ("E26B0A", "FFFFFF"),  # naranja
        ("70AD47", "FFFFFF"),  # verde
        ("FFC000", "000000"),  # amarillo
        ("4472C4", "FFFFFF"),  # azul
        ("FF0000", "FFFFFF"),  # rojo
        ("7030A0", "FFFFFF"),  # morado
        ("00B0F0", "000000"),  # azul claro
        ("92D050", "000000"),  # verde claro
    ]
    # Color claro para las sub-filas de cada sucursal
    COLORES_CLARO = [
        "BDD7EE",  # azul claro
        "FCE4D6",  # naranja claro
        "E2EFDA",  # verde claro
        "FFEB9C",  # amarillo claro
        "DDEEFF",  # azul claro 2
        "FFD7D7",  # rojo claro
        "E4D5F0",  # morado claro
        "CCECFF",  # azul muy claro
        "D7F0BF",  # verde muy claro
    ]

    suc_idx = 0
    for suc in SUCURSALES:
        df_s = todos_nuevos[todos_nuevos["_SUCURSAL"] == suc]
        if df_s.empty: continue

        color_bg, color_fg = COLORES_SUCURSAL[suc_idx % len(COLORES_SUCURSAL)]
        color_sub = COLORES_CLARO[suc_idx % len(COLORES_CLARO)]

        # Fila de sucursal con total
        ws.merge_cells(start_row=fila, start_column=1,
                       end_row=fila,   end_column=2)
        c = ws.cell(fila, 1, suc.upper())
        ws.cell(fila, 3, len(df_s))
        s(c, bg=color_bg, bold=True, color=color_fg, size=10)
        s(ws.cell(fila, 3), bg=color_bg, bold=True, color=color_fg,
          size=10, h="center")
        for j in [4, 5]:
            ws.cell(fila, j).fill = fill(color_bg)
        ws.row_dimensions[fila].height = 20
        fila += 1

        # Desglose por status (con color claro de la sucursal)
        nagg = (df_s.groupby("_STATUS").size()
                .reset_index(name="N")
                .sort_values("N", ascending=False))
        for _, row in nagg.iterrows():
            ws.cell(fila, 1, "")
            ws.cell(fila, 2, row["_STATUS"])
            ws.cell(fila, 3, int(row["N"]))
            for j in range(1, 6):
                ws.cell(fila, j).fill = fill(color_sub)
            ws.cell(fila, 2).font = font(size=10)
            ws.cell(fila, 2).alignment = aln(h="left")
            ws.cell(fila, 3).font = font(size=10)
            ws.cell(fila, 3).alignment = aln(h="center")
            ws.row_dimensions[fila].height = 18
            fila += 1

        suc_idx += 1

    # Fila TOTAL NUEVOS
    ws.merge_cells(start_row=fila, start_column=1,
                   end_row=fila,   end_column=2)
    ws.cell(fila, 1, "TOTAL NUEVOS")
    ws.cell(fila, 3, len(todos_nuevos))
    for j in range(1, 6):
        s(ws.cell(fila, j), bg=C["azul_marino"], bold=True,
          color="FFFFFF", size=11, h="center" if j == 3 else "left")
    ws.row_dimensions[fila].height = 22

    ws.column_dimensions["A"].width = 40; ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 38; ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A6"

# ══════════════════════════════════════════════════════════════
# H2 — NOTAS BLOQUEADOS
# ══════════════════════════════════════════════════════════════

def hacer_h2(wb, df_act_all):
    ws = wb.create_sheet("H2 - Notas Bloqueados")
    titulo(ws, f"RESUMEN DE NOTAS POR GRUPO  |  {FECHA_ACTUAL_LARGA}  |  "
               f"Todas las sucursales", 1, 6)
    totales_g = []
    fila = 3

    for tg, cbg, sts, etqs, lbl_tot, ctot in [
        ("GRUPO: OPERACIONES", "azul_medio", OPS_STATUS,
         ["Bloq Ops", "Info Act Ops", "Reg Completo", "Verif Ops", "TOTAL OPS"],
         "TOTAL GRUPO OPERACIONES", C["azul_medio"]),
        ("GRUPO: KYC-PLD", "verde_oscuro", KYC_STATUS,
         ["Bloq PLD", "Info Act PLD", "Por Validar PLD", "Verif PLD", "TOTAL KYC-PLD"],
         "TOTAL GRUPO KYC-PLD", C["verde_oscuro"]),
    ]:
        seccion(ws, tg, fila, 6, bg=cbg); fila += 1
        subs    = [cnt(df_act_all, st) for st in sts]
        tot_g   = sum(subs); totales_g.append(tot_g)
        bg_enc  = C["azul_claro"] if cbg == "azul_medio" else C["verde_claro"]
        col_enc = C["azul_marino"] if cbg == "azul_medio" else C["verde_oscuro"]

        ws.cell(fila, 1, "")
        for j, (et, sub) in enumerate(zip(etqs[:4], subs), 2):
            c = ws.cell(fila, j, f"{et}\n{sub:,}")
            s(c, bg=bg_enc, bold=True, color=col_enc, size=10, h="center", wrap=True)
        c = ws.cell(fila, 6, f"{etqs[4]}\n{tot_g:,}")
        s(c, bg=C["naranja"], bold=True, color="FFFFFF", size=10, h="center", wrap=True)
        ws.row_dimensions[fila].height = 28; fila += 1
        ws.row_dimensions[fila].height = 12; fila += 1

        for j, et in enumerate(["CATEGORÍA DE NOTA"] + etqs[:4] + [etqs[4]], 1):
            c = ws.cell(fila, j, et)
            s(c, bg=C["azul_marino"], bold=True, color="FFFFFF",
              size=10, h="center" if j > 1 else "left")
        ws.row_dimensions[fila].height = 24; fila += 1

        for idx, cat in enumerate(CATEGORIAS_NOTAS):
            conteos  = [int(((df_act_all["_STATUS"] == st) &
                             (df_act_all["_NOTA_CAT"] == cat)).sum()) for st in sts]
            tot_cat  = sum(conteos)
            bg       = bg_enc if idx % 2 == 0 else C["blanco"]
            ws.cell(fila, 1, cat)
            s(ws.cell(fila, 1), bg=bg, size=10)
            for j, v in enumerate(conteos, 2):
                c = ws.cell(fila, j, v if v else "-")
                s(c, bg=bg, size=10, h="center")
            c = ws.cell(fila, 6, tot_cat if tot_cat else "-")
            s(c, bg=bg, bold=True, size=10, h="center")
            ws.row_dimensions[fila].height = 18; fila += 1

        ws.cell(fila, 1, lbl_tot)
        for j, v in enumerate(subs, 2): ws.cell(fila, j, v)
        ws.cell(fila, 6, tot_g)
        for j in range(1, 7):
            s(ws.cell(fila, j), bg=ctot, bold=True, color="FFFFFF",
              size=10, h="center" if j > 1 else "left")
        ws.row_dimensions[fila].height = 22; fila += 2

    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    ws.cell(fila, 1, "TOTAL GLOBAL (OPERACIONES + KYC-PLD)")
    ws.cell(fila, 6, sum(totales_g))
    s(ws.cell(fila, 1), bg=C["naranja"], bold=True, color="FFFFFF", size=11)
    s(ws.cell(fila, 6), bg=C["naranja"], bold=True, color="FFFFFF",
      size=11, h="center")
    ws.row_dimensions[fila].height = 22
    ws.column_dimensions["A"].width = 38
    for l in ["B", "C", "D", "E", "F"]: ws.column_dimensions[l].width = 18
    ws.freeze_panes = "A6"

# ══════════════════════════════════════════════════════════════
# H3 — DATOS POR SUCURSAL (con sub-filas por nivel de lealtad)
# ══════════════════════════════════════════════════════════════

def hacer_h3(wb, df_act_all, datos_por_sucursal):
    ws = wb.create_sheet("H3 - Datos por Sucursal")
    titulo(ws, f"RESUMEN DE DATOS IMPORTANTES POR SUCURSAL  |  {FECHA_ACTUAL_LARGA}",
           1, 10)

    ws.merge_cells("A2:J2")
    c = ws.cell(2, 1, "ID Vencida: fecha ≤ 31/12/2025")
    s(c, bg=C["azul_medio"], color="FFFFFF", size=10)
    ws.row_dimensions[2].height = 16

    for ini, fin, txt in [(4,5,"ID VENCIDA (≤31/12/25)"),
                          (6,7,"AVISO DE PRIVACIDAD"),
                          (8,9,"HUELLA DIGITAL")]:
        ws.merge_cells(start_row=3, start_column=ini, end_row=3, end_column=fin)
        c = ws.cell(3, ini, txt)
        s(c, bg=C["azul_medio"], bold=True, color="FFFFFF", size=10, h="center")
    ws.row_dimensions[3].height = 18

    for j, h in enumerate(["SUCURSAL", "NIVEL LEALTAD", "TOTAL", "CANTIDAD", "%",
                            "CON AVISO (SÍ)", "SIN AVISO (NO)",
                            "CON HUELLA", "SIN HUELLA", "% HUELLA"], 1):
        c = ws.cell(4, j, h)
        s(c, bg=C["azul_marino"], bold=True, color="FFFFFF", size=10,
          h="center" if j > 2 else "left", wrap=True)
    ws.row_dimensions[4].height = 24

    fila = 5
    for suc, _, _ in datos_por_sucursal:
        df_suc = df_act_all[df_act_all["_SUCURSAL"] == suc]
        if df_suc.empty: continue

        t = len(df_suc)
        id_v = df_suc["_ID_VENC"].sum()
        av   = df_suc["_AVISO"].sum()
        hu   = df_suc["_HUELLA"].sum()

        # Fila resumen de sucursal
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
        c = ws.cell(fila, 1, suc.upper())
        s(c, bg=C["azul_claro"], bold=True, color=C["azul_marino"], size=10)
        for j, v in enumerate([t, int(id_v), pct(id_v, t),
                                int(av), int(t-av),
                                int(hu), int(t-hu), pct(hu, t)], 3):
            c = ws.cell(fila, j, v)
            s(c, bg=C["azul_claro"], bold=True, color=C["azul_marino"],
              size=10, h="center")
        ws.row_dimensions[fila].height = 20; fila += 1

        # Sub-filas por nivel
        datos_nv = (df_suc.groupby("_NIVEL")
                    .agg(TOTAL=("_ID","count"), ID_VENC=("_ID_VENC","sum"),
                         CON_AVISO=("_AVISO","sum"), CON_HUELLA=("_HUELLA","sum"))
                    .reset_index()
                    .sort_values("TOTAL", ascending=False)
                    .reset_index(drop=True))

        for idx2 in range(len(datos_nv)):
            row   = datos_nv.iloc[idx2]
            nivel = str(row["_NIVEL"])
            bg    = COLOR_NIVEL.get(nivel, C["blanco"])
            ws.cell(fila, 1, "")
            ws.cell(fila, 2, nivel)
            s(ws.cell(fila, 2), bg=bg, size=10)
            s(ws.cell(fila, 1), bg=bg)
            for j, v in enumerate([int(row.TOTAL), int(row.ID_VENC),
                                    pct(row.ID_VENC, row.TOTAL),
                                    int(row.CON_AVISO),
                                    int(row.TOTAL - row.CON_AVISO),
                                    int(row.CON_HUELLA),
                                    int(row.TOTAL - row.CON_HUELLA),
                                    pct(row.CON_HUELLA, row.TOTAL)], 3):
                c = ws.cell(fila, j, v)
                s(c, bg=bg, size=10, h="center")
            ws.row_dimensions[fila].height = 18; fila += 1

    # Total general
    t_all  = len(df_act_all)
    id_all = df_act_all["_ID_VENC"].sum()
    av_all = df_act_all["_AVISO"].sum()
    hu_all = df_act_all["_HUELLA"].sum()
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    ws.cell(fila, 1, "TOTAL GENERAL")
    s(ws.cell(fila, 1), bg=C["naranja"], bold=True, color="FFFFFF", size=10)
    for j, v in enumerate([t_all, int(id_all), pct(id_all, t_all),
                            int(av_all), int(t_all-av_all),
                            int(hu_all), int(t_all-hu_all),
                            pct(hu_all, t_all)], 3):
        c = ws.cell(fila, j, v)
        s(c, bg=C["naranja"], bold=True, color="FFFFFF", size=10, h="center")
    ws.row_dimensions[fila].height = 22

    ws.column_dimensions["A"].width = 16; ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    for l in ["D","E","F","G","H","I","J"]: ws.column_dimensions[l].width = 14
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════
# H4 — RESUMEN ESTATUS (global + por sucursal)
# ══════════════════════════════════════════════════════════════

def hacer_h4(wb, df_ant_all, df_act_all, datos_por_sucursal):
    ws = wb.create_sheet("H4 - Resumen Estatus")
    titulo(ws, f"RESUMEN DE ESTATUS  |  {FECHA_ANTERIOR_LARGA} vs {FECHA_ACTUAL_LARGA}",
           1, 4)

    seccion(ws, "RESUMEN GLOBAL", 3, 4)
    ws.row_dimensions[3].height = 22
    fila = bloque_estatus(ws, 4, "TOTAL BASE DE DATOS", df_ant_all, df_act_all)

    fila += 1
    seccion(ws, "DESGLOSE POR SUCURSAL", fila, 4); fila += 1

    for suc, df_ant_suc, df_act_suc in datos_por_sucursal:
        ws.merge_cells(start_row=fila, start_column=1,
                       end_row=fila,   end_column=4)
        c = ws.cell(fila, 1, f"  {suc.upper()}")
        s(c, bg=C["azul_claro"], bold=True, color=C["azul_marino"], size=11)
        ws.row_dimensions[fila].height = 22; fila += 1
        fila = bloque_estatus(ws, fila, f"Total {suc}", df_ant_suc, df_act_suc)
        fila += 1

    ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18; ws.column_dimensions["D"].width = 16
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════
# H5 — RESUMEN GLOBAL ESTADÍSTICO
# ══════════════════════════════════════════════════════════════

def hacer_h5(wb, df_ant_all, df_act_all, todos_cambios, todos_nuevos):
    ws = wb.create_sheet("H5 - Resumen Global")
    ta, tc = len(df_ant_all), len(df_act_all)

    titulo(ws, f"RESUMEN GLOBAL ESTADÍSTICO  |  Análisis PLD  |  "
               f"{FECHA_ANTERIOR_CORTA} vs {FECHA_ACTUAL_CORTA}", 1, 6, size=15)
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A2:F2")
    c = ws.cell(2, 1, f"Análisis PLD – {len(SUCURSALES)} sucursal(es)  |  "
                      f"{FECHA_ANTERIOR_LARGA} vs {FECHA_ACTUAL_LARGA}")
    s(c, bg=C["azul_medio"], color="FFFFFF", size=11)
    ws.row_dimensions[2].height = 18

    enc(ws, ["INDICADOR", FECHA_ANTERIOR_CORTA, "%",
             FECHA_ACTUAL_CORTA, "%", "VARIACIÓN"], 4)

    ops_a = sum(cnt(df_ant_all, st) for st in OPS_STATUS)
    ops_c = sum(cnt(df_act_all, st) for st in OPS_STATUS)
    kyc_a = sum(cnt(df_ant_all, st) for st in KYC_STATUS)
    kyc_c = sum(cnt(df_act_all, st) for st in KYC_STATUS)

    def fh(fila, lbl, va, vc, bg=None, bold=False, sec=False):
        if sec:
            ws.merge_cells(start_row=fila, start_column=1,
                           end_row=fila, end_column=6)
            c = ws.cell(fila, 1, lbl)
            s(c, bg=C["azul_medio"], bold=True, color="FFFFFF", size=11)
            ws.row_dimensions[fila].height = 22; return
        if not lbl:
            ws.row_dimensions[fila].height = 8; return
        bg_use = bg or (C["azul_claro"] if fila % 2 == 0 else C["blanco"])
        vr = var_str(vc - va)
        for j, v in enumerate([lbl, va, pct(va, ta), vc, pct(vc, tc), vr], 1):
            c = ws.cell(fila, j, v)
            if j == 6:
                var_cell(c, vr)
                c.font = Font(bold=bold, size=10, name="Arial")
            else:
                s(c, bg=bg_use, bold=bold, size=10,
                  h="left" if j == 1 else "center")
        ws.row_dimensions[fila].height = 22

    fh(5,  "── UNIVERSO TOTAL ────────────────────────", 0, 0, sec=True)
    fh(6,  "Total Clientes en Base", ta, tc)
    fh(7,  "Clientes Nuevos (entre cortes)", 0, len(todos_nuevos))
    fh(8,  "", 0, 0)
    fh(9,  "── ESTATUS ───────────────────────────────", 0, 0, sec=True)
    fh(10, "Registro Express",
       cnt(df_ant_all,"Registro express"), cnt(df_act_all,"Registro express"))
    fh(11, "Grupo Operaciones – SUBTOTAL", ops_a, ops_c, C["azul_claro"], True)
    fh(12, "  Bloqueado operaciones",
       cnt(df_ant_all,"Bloqueado operaciones"),
       cnt(df_act_all,"Bloqueado operaciones"))
    fh(13, "  Info actualizada operaciones",
       cnt(df_ant_all,"Info actualizada operaciones"),
       cnt(df_act_all,"Info actualizada operaciones"))
    fh(14, "  Registro Completo",
       cnt(df_ant_all,"Registro completo"), cnt(df_act_all,"Registro completo"))
    fh(15, "  Verificado operaciones",
       cnt(df_ant_all,"Verificado operaciones"),
       cnt(df_act_all,"Verificado operaciones"))
    fh(16, "Grupo KYC-PLD – SUBTOTAL", kyc_a, kyc_c, C["verde_claro"], True)
    fh(17, "  Bloqueado PLD",
       cnt(df_ant_all,"Bloqueado PLD"), cnt(df_act_all,"Bloqueado PLD"))
    fh(18, "  Info actualizada PLD",
       cnt(df_ant_all,"Info actualizada PLD"),
       cnt(df_act_all,"Info actualizada PLD"))
    fh(19, "  Por validar PLD",
       cnt(df_ant_all,"Por validar PLD"),
       cnt(df_act_all,"Por validar PLD"))
    fh(20, "  Verificado PLD",
       cnt(df_ant_all,"Verificado PLD"), cnt(df_act_all,"Verificado PLD"))
    fh(21, "", 0, 0)
    fh(22, "── DOCUMENTACIÓN ─────────────────────────", 0, 0, sec=True)
    fh(23, "ID Vencida (expiración ≤ 31/12/2025)",
       int(df_ant_all["_ID_VENC"].sum()), int(df_act_all["_ID_VENC"].sum()))
    fh(24, "Con Aviso de Privacidad Firmado (SÍ)",
       int(df_ant_all["_AVISO"].sum()), int(df_act_all["_AVISO"].sum()))
    fh(25, "Sin Aviso de Privacidad Firmado (NO)",
       ta-int(df_ant_all["_AVISO"].sum()), tc-int(df_act_all["_AVISO"].sum()))
    fh(26, "Con Huella Digital (SÍ)",
       int(df_ant_all["_HUELLA"].sum()), int(df_act_all["_HUELLA"].sum()))
    fh(27, "Sin Huella Digital (NO)",
       ta-int(df_ant_all["_HUELLA"].sum()), tc-int(df_act_all["_HUELLA"].sum()))
    fh(28, "", 0, 0)
    fh(29, "── MOVIMIENTOS ENTRE CORTES ──────────────", 0, 0, sec=True)
    fh(30, "Total Cambios de Estatus", 0, len(todos_cambios))

    trans_all = (todos_cambios
                 .groupby(["_STATUS_ANT", "_STATUS_ACT"]).size()
                 .reset_index(name="N").sort_values("N", ascending=False))
    for i, (_, row) in enumerate(trans_all.iterrows(), 31):
        fh(i, f"  {row['_STATUS_ANT'][:24]}  →  {row['_STATUS_ACT'][:24]}",
           0, int(row["N"]))

    ultima = 31 + len(trans_all) + 1
    ws.merge_cells(start_row=ultima, start_column=1, end_row=ultima, end_column=6)
    c = ws.cell(ultima, 1,
                "⚠  Nota: Las variaciones negativas en estatus de bloqueo "
                "indican mejora en la regularización de clientes.")
    s(c, bg=C["gris_claro"], size=10, color="595959")
    ws.row_dimensions[ultima].height = 20

    ws.column_dimensions["A"].width = 40; ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12; ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12; ws.column_dimensions["F"].width = 16
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════
# H6 — DETALLE REGISTRO EXPRESS
# ══════════════════════════════════════════════════════════════

def hacer_h6(wb, todos_cambios, df_act_all):
    ws  = wb.create_sheet("H6 - Detalle Reg Express")
    ncols = 6
    titulo(ws, f"DETALLE: CLIENTES EN REGISTRO EXPRESS QUE CAMBIARON DE ESTATUS  |  "
               f"{FECHA_ACTUAL_LARGA}", 1, ncols)

    reg_exp  = todos_cambios[todos_cambios["_STATUS_ANT"] == "Registro express"].copy()
    total_re = len(reg_exp)

    ws.merge_cells("A2:F2")
    c = ws.cell(2, 1, f"Total clientes: {total_re:,}   |   Corte: {FECHA_ACTUAL_LARGA}")
    s(c, bg=C["azul_medio"], bold=True, color="FFFFFF", size=11)
    ws.row_dimensions[2].height = 20

    dist     = reg_exp.groupby("_STATUS_ACT").size().sort_values(ascending=False)
    dist_txt = ("   Distribución:  " +
                "  |  ".join([f"{st}: {n:,}" for st, n in dist.items()]))
    ws.merge_cells("A3:F3")
    c = ws.cell(3, 1, dist_txt)
    s(c, bg=C["gris_claro"], size=10, color="595959")
    ws.row_dimensions[3].height = 16

    enc(ws, ["SUCURSAL", "PLAYER ID", "NOMBRE DEL CLIENTE",
             f"STATUS ANTERIOR\n({FECHA_ANTERIOR_CORTA})",
             f"STATUS NUEVO\n({FECHA_ACTUAL_CORTA})",
             "FECHA DE CAMBIO"], 5)
    ws.row_dimensions[5].height = 32
    fila = 6

    for suc in SUCURSALES:
        df_s = reg_exp[reg_exp["_SUCURSAL"] == suc]
        if df_s.empty: continue

        ws.merge_cells(start_row=fila, start_column=1,
                       end_row=fila,   end_column=ncols)
        c = ws.cell(fila, 1, f"  {suc.upper()}  —  {len(df_s):,} clientes")
        s(c, bg=C["azul_claro"], bold=True, color=C["azul_marino"], size=10)
        ws.row_dimensions[fila].height = 20; fila += 1

        for nst in df_s["_STATUS_ACT"].unique():
            df_st = df_s[df_s["_STATUS_ACT"] == nst]
            ws.merge_cells(start_row=fila, start_column=1,
                           end_row=fila,   end_column=ncols)
            c = ws.cell(fila, 1,
                        f"    → Nuevo estatus: {nst}  ({len(df_st):,})")
            s(c, bg=COLOR_STATUS.get(nst, C["gris_claro"]), size=10)
            ws.row_dimensions[fila].height = 18; fila += 1

            df_act_suc = df_act_all[df_act_all["_SUCURSAL"] == suc][
                ["_ID", "_NOMBRE", "_UPDATED"]]
            df_m = df_st.merge(df_act_suc, on="_ID", how="left",
                               suffixes=("", "_act"))

            for _, row in df_m.iterrows():
                fecha_str = (row["_UPDATED"].strftime("%d/%m/%Y %H:%M")
                             if pd.notna(row.get("_UPDATED")) else "")
                nombre = row.get("_NOMBRE", row.get("_NOMBRE_act", ""))
                for j, v in enumerate([suc, row["_ID"], nombre,
                                        row["_STATUS_ANT"], row["_STATUS_ACT"],
                                        fecha_str], 1):
                    c = ws.cell(fila, j, v); s(c, size=10)
                ws.row_dimensions[fila].height = 16; fila += 1
        fila += 1

    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 32; ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 28; ws.column_dimensions["F"].width = 20
    ws.freeze_panes = "A6"

# ══════════════════════════════════════════════════════════════
# H7 — DETALLE CLIENTES NUEVOS
# ══════════════════════════════════════════════════════════════

def hacer_h7(wb, todos_nuevos):
    ws    = wb.create_sheet("H7 - Detalle Clientes Nuevos")
    ncols = 4
    titulo(ws, f"DETALLE DE CLIENTES NUEVOS  |  {FECHA_ACTUAL_LARGA}", 1, ncols)

    ws.merge_cells("A2:D2")
    c = ws.cell(2, 1, f"Clientes en {FECHA_ACTUAL_CORTA} que no estaban en "
                      f"{FECHA_ANTERIOR_CORTA}  —  Total: {len(todos_nuevos):,}")
    s(c, bg=C["azul_medio"], bold=True, color="FFFFFF", size=11)
    ws.row_dimensions[2].height = 20

    dist_suc = todos_nuevos.groupby("_SUCURSAL").size()
    ws.merge_cells("A3:D3")
    c = ws.cell(3, 1, "  Por sucursal:  " +
                "  |  ".join([f"{k}: {v:,}" for k, v in dist_suc.items()]))
    s(c, bg=C["gris_claro"], size=10, color="595959")
    ws.row_dimensions[3].height = 16

    enc(ws, ["SUCURSAL", "PLAYER ID", "NOMBRE DEL CLIENTE",
             f"STATUS ACTUAL\n({FECHA_ACTUAL_CORTA})"], 5)
    ws.row_dimensions[5].height = 28
    fila = 6

    for suc in SUCURSALES:
        df_s = todos_nuevos[todos_nuevos["_SUCURSAL"] == suc]
        if df_s.empty: continue

        ws.merge_cells(start_row=fila, start_column=1,
                       end_row=fila,   end_column=ncols)
        c = ws.cell(fila, 1, f"  {suc.upper()}  —  {len(df_s):,} clientes nuevos")
        s(c, bg=C["azul_claro"], bold=True, color=C["azul_marino"], size=10)
        ws.row_dimensions[fila].height = 20; fila += 1

        for _, row in df_s.sort_values("_STATUS").iterrows():
            bg = COLOR_STATUS.get(row["_STATUS"], C["blanco"])
            for j, v in enumerate([suc, row["_ID"], row["_NOMBRE"],
                                    row["_STATUS"]], 1):
                c = ws.cell(fila, j, v); s(c, bg=bg, size=10)
            ws.row_dimensions[fila].height = 16; fila += 1
        fila += 1

    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 34; ws.column_dimensions["D"].width = 32
    ws.freeze_panes = "A6"

# ══════════════════════════════════════════════════════════════
# H8 — RESUMEN REGISTROS POR SEMANA
# ══════════════════════════════════════════════════════════════

def hacer_h8(wb, df_ant_all, df_act_all):
    ws    = wb.create_sheet("H8 - Resumen Registros Semana")
    ncols = 11
    titulo(ws, f"RESUMEN DE REGISTROS POR SEMANA  |  "
               f"{FECHA_ANTERIOR_LARGA} vs {FECHA_ACTUAL_LARGA}", 1, ncols)

    sem_ant_lbl = (f"{SEMANA_ANT_INICIO.strftime('%d/%m/%Y')} – "
                   f"{SEMANA_ANT_FIN.strftime('%d/%m/%Y')}")
    sem_nva_lbl = (f"{SEMANA_NVA_INICIO.strftime('%d/%m/%Y')} – "
                   f"{SEMANA_NVA_FIN.strftime('%d/%m/%Y')}")

    COLS_ST = ["Registro express", "Registro completo",
               "Info actualizada operaciones", "Verificado operaciones",
               "Bloqueado operaciones", "Por validar PLD",
               "Info actualizada PLD", "Verificado PLD", "Bloqueado PLD"]
    ETQS    = ["Reg express", "Reg completo", "Info Act Ops", "Verif Ops",
               "Bloq Ops", "Por Valid PLD", "Info Act PLD", "Verif PLD", "Bloq PLD"]
    headers = ["Sucursal", "Total"] + ETQS

    def filtrar(df, ini, fin):
        m = (df["_CREATED"].notna() &
             (df["_CREATED"].dt.normalize() >= ini) &
             (df["_CREATED"].dt.normalize() <= fin))
        return df[m]

    def tabla(ws, fila, df, ttl, lbl):
        seccion(ws, f"▶  {ttl}  –  {lbl}", fila, ncols); fila += 1
        enc(ws, headers, fila); fila += 1
        tots = {st: 0 for st in COLS_ST}; tg = 0
        for suc in SUCURSALES:
            df_s    = df[df["_SUCURSAL"] == suc]
            t       = len(df_s)
            conteos = [cnt(df_s, st) for st in COLS_ST]
            tg += t
            for st, n in zip(COLS_ST, conteos): tots[st] += n
            for j, v in enumerate([suc, t] +
                                   [v if v else "-" for v in conteos], 1):
                c = ws.cell(fila, j, v)
                s(c, size=10, h="left" if j == 1 else "center")
            ws.row_dimensions[fila].height = 18; fila += 1
        ws.cell(fila, 1, "Total")
        for j, v in enumerate([tg] + [tots[st] for st in COLS_ST], 2):
            ws.cell(fila, j, v)
        for j in range(1, ncols+1):
            s(ws.cell(fila, j), bg=C["azul_marino"], bold=True,
              color="FFFFFF", size=10, h="center" if j > 1 else "left")
        ws.row_dimensions[fila].height = 20; fila += 1
        return fila, tots, tg

    df_sa = filtrar(df_ant_all, SEMANA_ANT_INICIO, SEMANA_ANT_FIN)
    df_sn = filtrar(df_act_all, SEMANA_NVA_INICIO, SEMANA_NVA_FIN)

    fila = 3
    fila, ta, ga = tabla(ws, fila, df_sa, "SEMANA ANTERIOR", sem_ant_lbl)
    fila += 1
    fila, tn, gn = tabla(ws, fila, df_sn, "SEMANA NUEVA",    sem_nva_lbl)
    fila += 1

    # Variación
    seccion(ws, "▶  VARIACIÓN  (Semana Nueva – Semana Anterior)", fila, ncols)
    fila += 1
    enc(ws, headers, fila); fila += 1

    for suc in SUCURSALES:
        dsa = filtrar(df_ant_all[df_ant_all["_SUCURSAL"]==suc],
                      SEMANA_ANT_INICIO, SEMANA_ANT_FIN)
        dsn = filtrar(df_act_all[df_act_all["_SUCURSAL"]==suc],
                      SEMANA_NVA_INICIO, SEMANA_NVA_FIN)
        tv  = len(dsn) - len(dsa)
        cv  = [cnt(dsn, st) - cnt(dsa, st) for st in COLS_ST]
        for j, v in enumerate([suc, var_str(tv)] +
                               [var_str(v) if v != 0 else "-" for v in cv], 1):
            c = ws.cell(fila, j, v)
            if j > 1: var_cell(c, v)
            else:     s(c, size=10)
        ws.row_dimensions[fila].height = 18; fila += 1

    ws.cell(fila, 1, "Total")
    ws.cell(fila, 2, var_str(gn - ga))
    var_cell(ws.cell(fila, 2), var_str(gn - ga))
    for i, st in enumerate(COLS_ST):
        v = tn.get(st, 0) - ta.get(st, 0)
        c = ws.cell(fila, i+3, var_str(v) if v != 0 else "-")
        var_cell(c, var_str(v) if v != 0 else "-")
    s(ws.cell(fila, 1), bg=C["azul_marino"], bold=True,
      color="FFFFFF", size=10)
    ws.row_dimensions[fila].height = 20

    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 10
    for l in ["C","D","E","F","G","H","I","J","K"]:
        ws.column_dimensions[l].width = 14
    ws.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════
# H9 — DETALLE REGISTROS POR SEMANA
# ══════════════════════════════════════════════════════════════

def hacer_h9(wb, df_ant_all, df_act_all):
    ws    = wb.create_sheet("H9 - Detalle Registros Semana")
    ncols = 6
    sem_ant_lbl = (f"{SEMANA_ANT_INICIO.strftime('%d/%m/%Y')} – "
                   f"{SEMANA_ANT_FIN.strftime('%d/%m/%Y')}")
    sem_nva_lbl = (f"{SEMANA_NVA_INICIO.strftime('%d/%m/%Y')} – "
                   f"{SEMANA_NVA_FIN.strftime('%d/%m/%Y')}")

    titulo(ws, f"DETALLE DE REGISTROS POR SEMANA  |  "
               f"{FECHA_ANTERIOR_LARGA} vs {FECHA_ACTUAL_LARGA}", 1, ncols)
    ws.merge_cells("A2:F2")
    c = ws.cell(2, 1, f"Semana anterior: {sem_ant_lbl}   |   "
                      f"Semana nueva: {sem_nva_lbl}")
    s(c, bg=C["azul_medio"], bold=True, color="FFFFFF", size=11)
    ws.row_dimensions[2].height = 20

    enc(ws, ["SEMANA", "SUCURSAL", "PLAYER ID", "NOMBRE DEL CLIENTE",
             "STATUS ACTUAL",
             f"FECHA CREACIÓN\n(DATE_CREATED_NEW_RECORD)"], 4)
    ws.row_dimensions[4].height = 32
    fila = 5

    def filtrar(df, ini, fin):
        m = (df["_CREATED"].notna() &
             (df["_CREATED"].dt.normalize() >= ini) &
             (df["_CREATED"].dt.normalize() <= fin))
        return df[m]

    for lbl_sem, df_base, ini, fin in [
        (f"ANTERIOR\n{sem_ant_lbl}", df_ant_all, SEMANA_ANT_INICIO, SEMANA_ANT_FIN),
        (f"NUEVA\n{sem_nva_lbl}",    df_act_all, SEMANA_NVA_INICIO, SEMANA_NVA_FIN),
    ]:
        df_sem  = filtrar(df_base, ini, fin)
        cabecera = ("SEMANA ANTERIOR" if "ANTERIOR" in lbl_sem
                    else "SEMANA NUEVA")
        ws.merge_cells(start_row=fila, start_column=1,
                       end_row=fila,   end_column=ncols)
        c = ws.cell(fila, 1,
                    f"▶  {cabecera}  "
                    f"({ini.strftime('%d/%m/%Y')} – {fin.strftime('%d/%m/%Y')})"
                    f"  —  {len(df_sem):,} registros")
        s(c, bg=C["azul_marino"], bold=True, color="FFFFFF", size=11)
        ws.row_dimensions[fila].height = 22; fila += 1

        for suc in SUCURSALES:
            df_suc = df_sem[df_sem["_SUCURSAL"] == suc]
            if df_suc.empty: continue
            ws.merge_cells(start_row=fila, start_column=1,
                           end_row=fila,   end_column=ncols)
            c = ws.cell(fila, 1,
                        f"    {suc.upper()}  —  {len(df_suc):,} registros")
            s(c, bg=C["azul_claro"], bold=True, color=C["azul_marino"], size=10)
            ws.row_dimensions[fila].height = 18; fila += 1

            for _, row in df_suc.sort_values("_CREATED").iterrows():
                fecha_str = (row["_CREATED"].strftime("%d/%m/%Y %H:%M")
                             if pd.notna(row["_CREATED"]) else "")
                bg = COLOR_STATUS.get(row["_STATUS"], C["blanco"])
                for j, v in enumerate([lbl_sem, suc, row["_ID"], row["_NOMBRE"],
                                        row["_STATUS"], fecha_str], 1):
                    c = ws.cell(fila, j, v)
                    s(c, bg=bg, size=10, wrap=(j == 1))
                ws.row_dimensions[fila].height = 16; fila += 1
        fila += 1

    ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 38; ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 28; ws.column_dimensions["F"].width = 20
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════
# H10 — RESUMEN YTD 2026
# ══════════════════════════════════════════════════════════════

def hacer_h10(wb, df_ant_all, df_act_all):
    ws    = wb.create_sheet("H10 - Resumen YTD 2026")
    ncols = 11
    titulo(ws, f"RESUMEN YTD 2026  |  "
               f"Clientes creados desde {YTD_INICIO.strftime('%d/%m/%Y')}",
           1, ncols)
    ws.merge_cells("A2:K2")
    c = ws.cell(2, 1, f"Período: {YTD_INICIO.strftime('%d/%m/%Y')} – "
                      f"{FECHA_ACTUAL_LARGA}")
    s(c, bg=C["azul_medio"], bold=True, color="FFFFFF", size=11)
    ws.row_dimensions[2].height = 20

    COLS_ST = ["Registro express", "Registro completo",
               "Info actualizada operaciones", "Verificado operaciones",
               "Bloqueado operaciones", "Por validar PLD",
               "Info actualizada PLD", "Verificado PLD", "Bloqueado PLD"]
    ETQS    = ["Reg express", "Reg completo", "Info Act Ops", "Verif Ops",
               "Bloq Ops", "Por Valid PLD", "Info Act PLD", "Verif PLD", "Bloq PLD"]
    headers = ["Sucursal", "Total"] + ETQS

    def filtrar_ytd(df):
        return df[df["_CREATED"].notna() &
                  (df["_CREATED"].dt.normalize() >= YTD_INICIO)]

    def tabla_ytd(ws, fila, df, ttl):
        seccion(ws, f"▶  {ttl}", fila, ncols); fila += 1
        enc(ws, headers, fila); fila += 1
        tots = {st: 0 for st in COLS_ST}; tg = 0
        for suc in SUCURSALES:
            df_s    = df[df["_SUCURSAL"] == suc]
            t       = len(df_s)
            conteos = [cnt(df_s, st) for st in COLS_ST]
            tg += t
            for st, n in zip(COLS_ST, conteos): tots[st] += n
            for j, v in enumerate([suc, t] +
                                   [v if v else "-" for v in conteos], 1):
                c = ws.cell(fila, j, v)
                s(c, size=10, h="left" if j == 1 else "center")
            ws.row_dimensions[fila].height = 18; fila += 1
        ws.cell(fila, 1, "Total")
        for j, v in enumerate([tg] + [tots[st] for st in COLS_ST], 2):
            ws.cell(fila, j, v)
        for j in range(1, ncols+1):
            s(ws.cell(fila, j), bg=C["azul_marino"], bold=True,
              color="FFFFFF", size=10, h="center" if j > 1 else "left")
        ws.row_dimensions[fila].height = 20; fila += 1
        return fila, tots, tg

    fila = 4
    fila, ta_y, ga_y = tabla_ytd(ws, fila, filtrar_ytd(df_ant_all),
                                  f"BASE ANTERIOR  –  {FECHA_ANTERIOR_CORTA}")
    fila += 1
    fila, tn_y, gn_y = tabla_ytd(ws, fila, filtrar_ytd(df_act_all),
                                  f"BASE NUEVA  –  {FECHA_ACTUAL_CORTA}")
    fila += 1

    seccion(ws, "▶  VARIACIÓN  (Base Nueva – Base Anterior)", fila, ncols)
    fila += 1; enc(ws, headers, fila); fila += 1

    for suc in SUCURSALES:
        da = filtrar_ytd(df_ant_all[df_ant_all["_SUCURSAL"] == suc])
        dn = filtrar_ytd(df_act_all[df_act_all["_SUCURSAL"] == suc])
        tv = len(dn) - len(da)
        cv = [cnt(dn, st) - cnt(da, st) for st in COLS_ST]
        for j, v in enumerate([suc, var_str(tv)] +
                               [var_str(v) if v != 0 else "-" for v in cv], 1):
            c = ws.cell(fila, j, v)
            if j > 1: var_cell(c, v)
            else:     s(c, size=10)
        ws.row_dimensions[fila].height = 18; fila += 1

    ws.cell(fila, 1, "Total")
    ws.cell(fila, 2, var_str(gn_y - ga_y))
    var_cell(ws.cell(fila, 2), var_str(gn_y - ga_y))
    for i, st in enumerate(COLS_ST):
        v = tn_y.get(st, 0) - ta_y.get(st, 0)
        c = ws.cell(fila, i+3, var_str(v) if v != 0 else "-")
        var_cell(c, var_str(v) if v != 0 else "-")
    s(ws.cell(fila, 1), bg=C["azul_marino"], bold=True,
      color="FFFFFF", size=10)
    ws.row_dimensions[fila].height = 20

    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 10
    for l in ["C","D","E","F","G","H","I","J","K"]:
        ws.column_dimensions[l].width = 14
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════
# H11 — DETALLE YTD 2026
# ══════════════════════════════════════════════════════════════

def hacer_h11(wb, df_act_all):
    ws    = wb.create_sheet("H11 - Detalle YTD 2026")
    ncols = 5
    titulo(ws, f"DETALLE YTD 2026  |  "
               f"Clientes creados desde {YTD_INICIO.strftime('%d/%m/%Y')}  |  "
               f"{FECHA_ACTUAL_LARGA}", 1, ncols)

    df_ytd = df_act_all[
        df_act_all["_CREATED"].notna() &
        (df_act_all["_CREATED"].dt.normalize() >= YTD_INICIO)
    ].copy()

    ws.merge_cells("A2:E2")
    c = ws.cell(2, 1, f"Base: {FECHA_ACTUAL_CORTA}  —  "
                      f"Total: {len(df_ytd):,} clientes")
    s(c, bg=C["azul_medio"], bold=True, color="FFFFFF", size=11)
    ws.row_dimensions[2].height = 20

    dist2 = df_ytd.groupby("_SUCURSAL").size()
    ws.merge_cells("A3:E3")
    c = ws.cell(3, 1, "  Por sucursal:  " +
                "  |  ".join([f"{k}: {v:,}" for k, v in dist2.items()]))
    s(c, bg=C["gris_claro"], size=10, color="595959")
    ws.row_dimensions[3].height = 16

    enc(ws, ["SUCURSAL", "PLAYER ID", "NOMBRE DEL CLIENTE",
             f"STATUS ACTUAL\n({FECHA_ACTUAL_CORTA})",
             f"FECHA CREACIÓN\n(DATE_CREATED_NEW_RECORD)"], 5)
    ws.row_dimensions[5].height = 32
    fila = 6

    for suc in SUCURSALES:
        df_suc = df_ytd[df_ytd["_SUCURSAL"] == suc]
        if df_suc.empty: continue

        ws.merge_cells(start_row=fila, start_column=1,
                       end_row=fila,   end_column=ncols)
        c = ws.cell(fila, 1, f"  {suc.upper()}  —  {len(df_suc):,} clientes")
        s(c, bg=C["azul_claro"], bold=True, color=C["azul_marino"], size=10)
        ws.row_dimensions[fila].height = 20; fila += 1

        for st in TODOS_STATUS:
            df_st = df_suc[df_suc["_STATUS"] == st]
            if df_st.empty: continue
            ws.merge_cells(start_row=fila, start_column=1,
                           end_row=fila,   end_column=ncols)
            c = ws.cell(fila, 1, f"    → {st}  ({len(df_st):,})")
            s(c, bg=COLOR_STATUS.get(st, C["gris_claro"]), size=10)
            ws.row_dimensions[fila].height = 18; fila += 1

            for _, row in df_st.sort_values("_CREATED").iterrows():
                fecha_str = (row["_CREATED"].strftime("%d/%m/%Y %H:%M")
                             if pd.notna(row["_CREATED"]) else "")
                bg = COLOR_STATUS.get(row["_STATUS"], C["blanco"])
                for j, v in enumerate([suc, row["_ID"], row["_NOMBRE"],
                                        row["_STATUS"], fecha_str], 1):
                    c = ws.cell(fila, j, v); s(c, bg=bg, size=10)
                ws.row_dimensions[fila].height = 16; fila += 1
        fila += 1

    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 34; ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 22
    ws.freeze_panes = "A6"

# ══════════════════════════════════════════════════════════════
# PROCESO PRINCIPAL
# ══════════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════════
# GENERAR datos.json (para actualización automática de la página)
# ══════════════════════════════════════════════════════════════

def generar_json(df_ant_all, df_act_all, datos_por_sucursal,
                 todos_cambios, todos_nuevos):
    ta = len(df_ant_all); tc = len(df_act_all)
    def c(df, st): return int((df["_STATUS"] == st).sum())

    sucursales = []
    for suc, df_ant, df_act in datos_por_sucursal:
        ta_s = len(df_ant); tc_s = len(df_act)
        nuevos_suc = len(todos_nuevos[todos_nuevos["_SUCURSAL"] == suc]) if "_SUCURSAL" in todos_nuevos.columns else 0
        cambios_suc = len(todos_cambios[todos_cambios["_SUCURSAL"] == suc]) if "_SUCURSAL" in todos_cambios.columns else 0
        sucursales.append({
            "nombre": suc, "total_ant": ta_s, "total_act": tc_s,
            "delta": tc_s - ta_s, "nuevos": int(nuevos_suc), "cambios": int(cambios_suc),
            "re_act":        c(df_act,"Registro express"),
            "bloq_ops":      c(df_act,"Bloqueado operaciones"),
            "info_act_ops":  c(df_act,"Info actualizada operaciones"),
            "reg_completo":  c(df_act,"Registro completo"),
            "verif_ops":     c(df_act,"Verificado operaciones"),
            "bloq_kyc":      c(df_act,"Bloqueado PLD"),
            "info_act_kyc":  c(df_act,"Info actualizada PLD"),
            "pend_kyc":      c(df_act,"Por validar PLD"),
            "verif_kyc":     c(df_act,"Verificado PLD"),
            "id_venc_act": int(df_act["_ID_VENC"].sum()),
            "aviso_act":   int(df_act["_AVISO"].sum()),
            "huella_act":  int(df_act["_HUELLA"].sum()),
            "pct_id_venc": round(df_act["_ID_VENC"].sum() / tc_s * 100, 1) if tc_s else 0,
            "pct_aviso":   round(df_act["_AVISO"].sum()   / tc_s * 100, 1) if tc_s else 0,
            "pct_huella":  round(df_act["_HUELLA"].sum()  / tc_s * 100, 1) if tc_s else 0,
            "pct_verif_kyc": round(c(df_act,"Verificado PLD") / tc_s * 100, 1) if tc_s else 0,
            # Reporte de operaciones (clientes creados desde 01/01/2026 hasta fecha del corte actual)
            "ops_reg_completo":   int(df_act[df_act["_CREATED"].notna() & (df_act["_CREATED"].dt.normalize() >= YTD_INICIO) & (df_act["_STATUS"] == "Registro completo")].shape[0]),
            "ops_verif_ops":      int(df_act[df_act["_CREATED"].notna() & (df_act["_CREATED"].dt.normalize() >= YTD_INICIO) & (df_act["_STATUS"] == "Verificado operaciones")].shape[0]),
            "ops_bloq_ops":       int(df_act[df_act["_CREATED"].notna() & (df_act["_CREATED"].dt.normalize() >= YTD_INICIO) & (df_act["_STATUS"] == "Bloqueado operaciones")].shape[0]),
            "ops_info_act_ops":   int(df_act[df_act["_CREATED"].notna() & (df_act["_CREATED"].dt.normalize() >= YTD_INICIO) & (df_act["_STATUS"] == "Info actualizada operaciones")].shape[0]),
            "ops_re":             int(df_act[df_act["_CREATED"].notna() & (df_act["_CREATED"].dt.normalize() >= YTD_INICIO) & (df_act["_STATUS"] == "Registro express")].shape[0]),
            # KYC-PLD YTD
            "ops_bloq_kyc":       int(df_act[df_act["_CREATED"].notna() & (df_act["_CREATED"].dt.normalize() >= YTD_INICIO) & (df_act["_STATUS"] == "Bloqueado PLD")].shape[0]),
            "ops_info_act_kyc":   int(df_act[df_act["_CREATED"].notna() & (df_act["_CREATED"].dt.normalize() >= YTD_INICIO) & (df_act["_STATUS"] == "Info actualizada PLD")].shape[0]),
            "ops_pend_kyc":       int(df_act[df_act["_CREATED"].notna() & (df_act["_CREATED"].dt.normalize() >= YTD_INICIO) & (df_act["_STATUS"] == "Por validar PLD")].shape[0]),
            "ops_verif_kyc":      int(df_act[df_act["_CREATED"].notna() & (df_act["_CREATED"].dt.normalize() >= YTD_INICIO) & (df_act["_STATUS"] == "Verificado PLD")].shape[0]),
        })

    top_trans = []
    if len(todos_cambios) > 0 and "_STATUS_ANT" in todos_cambios.columns:
        top_trans = (todos_cambios
                     .groupby(["_STATUS_ANT","_STATUS_ACT"]).size()
                     .reset_index(name="N")
                     .sort_values("N", ascending=False)
                     .to_dict("records"))

    datos = {
        "meta": {
            "generado":             datetime.now().strftime("%d/%m/%Y %H:%M"),
            "fecha_anterior_corta": FECHA_ANTERIOR_CORTA,
            "fecha_actual_corta":   FECHA_ACTUAL_CORTA,
            "fecha_anterior_larga": FECHA_ANTERIOR_LARGA,
            "fecha_actual_larga":   FECHA_ACTUAL_LARGA,
            "num_sucursales":       len(datos_por_sucursal),
        },
        "kpis": {
            "total_ant": ta, "total_act": tc, "delta_total": tc - ta,
            "clientes_nuevos": len(todos_nuevos), "cambios_status": len(todos_cambios),
            "ops_ant": sum(c(df_ant_all,s) for s in OPS_STATUS),
            "ops_act": sum(c(df_act_all,s) for s in OPS_STATUS),
            "kyc_ant": sum(c(df_ant_all,s) for s in KYC_STATUS),
            "kyc_act": sum(c(df_act_all,s) for s in KYC_STATUS),
            "re_ant":  c(df_ant_all,"Registro express"),
            "re_act":  c(df_act_all,"Registro express"),
            "bloq_ops_ant":      c(df_ant_all,"Bloqueado operaciones"),
            "bloq_ops_act":      c(df_act_all,"Bloqueado operaciones"),
            "info_act_ops_ant":  c(df_ant_all,"Info actualizada operaciones"),
            "info_act_ops_act":  c(df_act_all,"Info actualizada operaciones"),
            "reg_completo_ant":  c(df_ant_all,"Registro completo"),
            "reg_completo_act":  c(df_act_all,"Registro completo"),
            "verif_ops_ant":     c(df_ant_all,"Verificado operaciones"),
            "verif_ops_act":     c(df_act_all,"Verificado operaciones"),
            "bloq_kyc_ant":      c(df_ant_all,"Bloqueado PLD"),
            "bloq_kyc_act":      c(df_act_all,"Bloqueado PLD"),
            "info_act_kyc_ant":  c(df_ant_all,"Info actualizada PLD"),
            "info_act_kyc_act":  c(df_act_all,"Info actualizada PLD"),
            "pend_kyc_ant":      c(df_ant_all,"Por validar PLD"),
            "pend_kyc_act":      c(df_act_all,"Por validar PLD"),
            "verif_kyc_ant":     c(df_ant_all,"Verificado PLD"),
            "verif_kyc_act":     c(df_act_all,"Verificado PLD"),
            "id_venc_ant": int(df_ant_all["_ID_VENC"].sum()),
            "id_venc_act": int(df_act_all["_ID_VENC"].sum()),
            "aviso_ant": int(df_ant_all["_AVISO"].sum()),
            "aviso_act": int(df_act_all["_AVISO"].sum()),
            "huella_ant": int(df_ant_all["_HUELLA"].sum()),
            "huella_act": int(df_act_all["_HUELLA"].sum()),
        },
        "sucursales": sucursales,
        "top_transiciones": top_trans,
        "detalle_cambios": [
            {
                "sucursal": row["_SUCURSAL"],
                "id":       row["_ID"],
                "ant":      row["_STATUS_ANT"],
                "act":      row["_STATUS_ACT"],
            }
            for _, row in todos_cambios.iterrows()
        ] if len(todos_cambios) < 100000 else [],
        "reporte_operaciones": {
            "fecha_inicio": YTD_INICIO.strftime("%d/%m/%Y"),
            "fecha_fin":    SEMANA_NVA_FIN.strftime("%d/%m/%Y"),
            "sucursales": [
                {
                    "nombre":       s["nombre"],
                    "total":        (s["ops_reg_completo"] + s["ops_verif_ops"] +
                                     s["ops_bloq_ops"] + s["ops_info_act_ops"] + s["ops_re"]),
                    "reg_completo": s["ops_reg_completo"],
                    "verif_ops":    s["ops_verif_ops"],
                    "bloq_ops":     s["ops_bloq_ops"],
                    "info_act_ops": s["ops_info_act_ops"],
                    "re":           s["ops_re"],
                    # KYC-PLD
                    "kyc_bloq":     s["ops_bloq_kyc"],
                    "kyc_info_act": s["ops_info_act_kyc"],
                    "kyc_pend":     s["ops_pend_kyc"],
                    "kyc_verif":    s["ops_verif_kyc"],
                    "kyc_total":    (s["ops_bloq_kyc"] + s["ops_info_act_kyc"] +
                                     s["ops_pend_kyc"] + s["ops_verif_kyc"]),
                }
                for s in sucursales
            ]
        },
    }

    # Guardar datos.json principal (compatibilidad)
    ruta = CARPETA_SALIDA / "datos.json"
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)
    print(f"  ✅ datos.json → {ruta}")

    # Guardar copia en carpeta datos/ para historial
    carpeta_datos = CARPETA_SALIDA / "datos"
    carpeta_datos.mkdir(exist_ok=True)
    nombre_corte = f"{FECHA_ANTERIOR_CORTA}_vs_{FECHA_ACTUAL_CORTA}.json"
    ruta_hist = carpeta_datos / nombre_corte
    with open(ruta_hist, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)
    print(f"  ✅ Historial → datos/{nombre_corte}")

    # Actualizar cortes.json
    ruta_cortes = CARPETA_SALIDA / "cortes.json"
    etiqueta = f"{FECHA_ANTERIOR_CORTA} vs {FECHA_ACTUAL_CORTA}"
    archivo  = f"datos/{nombre_corte}"
    entrada  = {"etiqueta": etiqueta, "archivo": archivo,
                "generado": datetime.now().strftime("%d/%m/%Y %H:%M")}

    # Leer lista existente o crear nueva
    if ruta_cortes.exists():
        with open(ruta_cortes) as f:
            cortes = json.load(f)
        # Actualizar si ya existe, agregar si es nuevo
        exists = [i for i, c in enumerate(cortes) if c["archivo"] == archivo]
        if exists:
            cortes[exists[0]] = entrada
        else:
            cortes.insert(0, entrada)  # más reciente primero
    else:
        cortes = [entrada]

    with open(ruta_cortes, "w", encoding="utf-8") as f:
        json.dump(cortes, f, ensure_ascii=False, indent=2)
    print(f"  ✅ cortes.json actualizado ({len(cortes)} corte(s))")


# ══════════════════════════════════════════════════════════════
# EXCEL DETALLE DE CAMBIOS DE ESTATUS
# ══════════════════════════════════════════════════════════════


def generar_base_unificada(df_act_all):
    """Genera un Excel con la base completa de clientes al corte actual,
    todas las sucursales en una sola hoja, lista para filtrar y compartir."""
    from openpyxl import Workbook

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Base Unificada")

    # Título
    titulo(ws, f"BASE UNIFICADA · {FECHA_ACTUAL_LARGA.upper()}  |  TODAS LAS SUCURSALES", 1, 9)

    # Encabezados
    cols_enc = [
        "SUCURSAL", "PLAYER ID", "NOMBRE", "ESTATUS",
        "NOTAS", "NIVEL", "AVISO FIRMADO", "HUELLA", "VENC. IFE"
    ]
    enc(ws, cols_enc, 2)
    ws.row_dimensions[2].height = 28
    ws.auto_filter.ref = f"A2:I{len(df_act_all)+2}"
    ws.freeze_panes = "A3"

    # Ordenar por sucursal y estatus
    df = df_act_all.sort_values(["_SUCURSAL", "_STATUS"]).reset_index(drop=True)

    for i, (_, row) in enumerate(df.iterrows(), 3):
        bg = COLOR_STATUS.get(row["_STATUS"], C["blanco"])

        # Formatear vencimiento
        exp_val = ""
        if pd.notna(row.get("_EXP")):
            exp_val = row["_EXP"].strftime("%d/%m/%Y")

        vals = [
            row.get("_SUCURSAL", ""),
            row["_ID"],
            row.get("_NOMBRE", ""),
            row.get("_STATUS", ""),
            row.get("_NOTAS", ""),
            row.get("_NIVEL", ""),
            "Sí" if row.get("_AVISO") else "No",
            "Sí" if row.get("_HUELLA") else "No",
            exp_val,
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v)
            s(c, bg=bg if j == 4 else C["blanco"], size=10)
        ws.row_dimensions[i].height = 15

    # Anchos de columna
    anchos = [14, 36, 28, 30, 32, 16, 14, 10, 14]
    for col, ancho in zip("ABCDEFGHI", anchos):
        ws.column_dimensions[col].width = ancho

    # Hoja resumen por sucursal
    ws2 = wb.create_sheet("Resumen por Sucursal")
    titulo(ws2, f"RESUMEN POR SUCURSAL · {FECHA_ACTUAL_LARGA.upper()}", 1, len(KYC_STATUS)+len(OPS_STATUS)+3)
    cols_res = ["SUCURSAL", "TOTAL"] + OPS_STATUS + KYC_STATUS
    enc(ws2, cols_res, 2)
    ws2.row_dimensions[2].height = 28
    ws2.freeze_panes = "A3"

    sucursales = df["_SUCURSAL"].unique()
    for i, suc in enumerate(sorted(sucursales), 3):
        df_s = df[df["_SUCURSAL"] == suc]
        vals = [suc, len(df_s)] +                [cnt(df_s, st) for st in OPS_STATUS] +                [cnt(df_s, st) for st in KYC_STATUS]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(i, j, v)
            s(c, size=10, h="center" if j > 1 else "left")
        ws2.row_dimensions[i].height = 18

    # Fila total
    i_tot = len(sucursales) + 3
    vals_tot = ["TOTAL", len(df)] +                [cnt(df, st) for st in OPS_STATUS] +                [cnt(df, st) for st in KYC_STATUS]
    for j, v in enumerate(vals_tot, 1):
        c = ws2.cell(i_tot, j, v)
        s(c, bg=C["azul_marino"], bold=True, color="FFFFFF",
          size=10, h="center" if j > 1 else "left")
    ws2.row_dimensions[i_tot].height = 22

    for col, ancho in zip("ABCDEFGHIJ", [14,10,18,18,18,18,18,18,18,18]):
        ws2.column_dimensions[col].width = ancho

    # Guardar
    nombre = f"Base_Unificada_{FECHA_ACTUAL_CORTA}.xlsx"
    ruta   = CARPETA_SALIDA / nombre
    wb.save(ruta)
    kb = ruta.stat().st_size // 1024
    print(f"  ✅ {nombre}  ({kb} KB)")
    return ruta


def generar_detalle_cambios(todos_cambios, todos_nuevos, df_act_all):
    """Genera un Excel con el detalle completo de cambios, fácil de filtrar."""
    from openpyxl import Workbook

    wb = Workbook(); wb.remove(wb.active)

    # ── Hoja 1: Todos los cambios ordenados ─────────────────
    ws1 = wb.create_sheet("Cambios de Estatus")
    titulo(ws1, f"DETALLE DE CAMBIOS DE ESTATUS  |  "
                f"{FECHA_ANTERIOR_CORTA} vs {FECHA_ACTUAL_CORTA}", 1, 7)

    enc(ws1, ["SUCURSAL", "PLAYER ID",
              f"STATUS ANTERIOR ({FECHA_ANTERIOR_CORTA})",
              f"STATUS NUEVO ({FECHA_ACTUAL_CORTA})"], 2)
    ws1.row_dimensions[2].height = 28
    ws1.auto_filter.ref = f"A2:D{len(todos_cambios)+2}"

    # Ordenar por STATUS_ANT, STATUS_ACT, SUCURSAL
    df_m = todos_cambios.sort_values(
        ["_STATUS_ANT","_STATUS_ACT","_SUCURSAL"]).reset_index(drop=True)

    for i, (_, row) in enumerate(df_m.iterrows(), 3):
        bg_ant = COLOR_STATUS.get(row.get("_STATUS_ANT",""), C["blanco"])
        bg_act = COLOR_STATUS.get(row.get("_STATUS_ACT",""), C["blanco"])
        vals = [row.get("_SUCURSAL",""), row["_ID"],
                row.get("_STATUS_ANT",""), row.get("_STATUS_ACT","")]
        for j, v in enumerate(vals, 1):
            c = ws1.cell(i, j, v)
            if j == 3: s(c, bg=bg_ant, size=10)
            elif j == 4: s(c, bg=bg_act, size=10)
            else: s(c, size=10)
        ws1.row_dimensions[i].height = 16

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 38
    ws1.column_dimensions["C"].width = 34
    ws1.column_dimensions["D"].width = 34
    ws1.freeze_panes = "A3"

    # ── Hoja 2: Resumen por transición ──────────────────────
    ws2 = wb.create_sheet("Resumen por Transición")
    titulo(ws2, f"RESUMEN DE TRANSICIONES  |  "
                f"{FECHA_ANTERIOR_CORTA} vs {FECHA_ACTUAL_CORTA}", 1, 4)

    enc(ws2, [f"STATUS ANTERIOR ({FECHA_ANTERIOR_CORTA})",
              f"STATUS NUEVO ({FECHA_ACTUAL_CORTA})",
              "CANTIDAD", "% DEL TOTAL"], 2)
    ws2.row_dimensions[2].height = 28
    ws2.auto_filter.ref = f"A2:D{len(todos_cambios)+2}"

    total_c = len(todos_cambios)
    trans = (todos_cambios.groupby(["_STATUS_ANT","_STATUS_ACT"]).size()
             .reset_index(name="N").sort_values("N", ascending=False)
             .reset_index(drop=True))

    for i, (_, row) in enumerate(trans.iterrows(), 3):
        bg = COLOR_STATUS.get(row["_STATUS_ANT"], C["blanco"])
        pct_v = f"{row['N']/total_c*100:.1f}%" if total_c else "0.0%"
        for j, v in enumerate([row["_STATUS_ANT"], row["_STATUS_ACT"],
                                int(row["N"]), pct_v], 1):
            c = ws2.cell(i, j, v)
            s(c, bg=bg, size=10, h="center" if j in [3,4] else "left")
        ws2.row_dimensions[i].height = 18

    # Fila total
    ws2.cell(i+1, 1, "TOTAL CAMBIOS"); ws2.cell(i+1, 3, total_c); ws2.cell(i+1, 4, "100.0%")
    for j in range(1, 5):
        s(ws2.cell(i+1, j), bg=C["azul_marino"], bold=True, color="FFFFFF",
          size=10, h="center" if j in [3,4] else "left")
    ws2.row_dimensions[i+1].height = 22

    ws2.column_dimensions["A"].width = 38; ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 14; ws2.column_dimensions["D"].width = 14
    ws2.freeze_panes = "A3"

    # Guardar
    nombre = f"Detalle_Cambios_Estatus_{FECHA_ANTERIOR_CORTA}_vs_{FECHA_ACTUAL_CORTA}.xlsx"
    ruta   = CARPETA_SALIDA / nombre
    wb.save(ruta)
    kb = ruta.stat().st_size // 1024
    print(f"  ✅ {nombre}  ({kb} KB)")
    return ruta



# ══════════════════════════════════════════════════════════════
# GENERAR index.html CON DATOS EMBEBIDOS
# ══════════════════════════════════════════════════════════════


def _generar_html_contenido(**kw):
    """Genera el contenido del index.html reemplazando valores en el HTML base."""
    import json as _j

    fa = kw["fa_corta"]; fc = kw["fc_corta"]
    fal = kw["fa_larga"]; fcl = kw["fc_larga"]
    k   = kw["k"];  tc = kw["tc"]
    fmt = lambda n: f"{int(n):,}"

    def _kpi_delta_pos(val, ant):
        delta = val - ant
        return f'<div class="kpi-delta d-pos">▲ +{fmt(delta)} vs anterior</div>'
    def _kpi_delta_neg(val, ant):
        delta = val - ant
        sign = "▲" if delta > 0 else "▼"
        return f'<div class="kpi-delta d-neg">{sign} {fmt(delta)} vs anterior</div>'

    pA = kw["pAviso"]; pSA = kw["pSinAv"]
    pH = kw["pHuella"]; pSH = kw["pSinHu"]
    pV = kw["pIdVenc"]
    gen = kw["generado"]
    ns  = kw["num_suc"]

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Reportes PLD — CECOM</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
:root{{--azul:#1F3864;--azul2:#2E75B6;--verde:#375623;--naranja:#E26B0A;--bg:#F4F6FA;--card:#fff;--texto:#1a1a2e;--muted:#6b7280;--border:#e5e7eb;--pos:#16a34a;--neg:#dc2626;--pos-bg:#dcfce7;--neg-bg:#fee2e2}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--texto)}}
header{{background:var(--azul);color:#fff;padding:0 2rem;position:sticky;top:0;z-index:100;box-shadow:0 2px 12px rgba(31,56,100,.4)}}
.hinner{{max-width:1200px;margin:0 auto;display:flex;align-items:center;justify-content:space-between;height:64px;gap:12px;flex-wrap:wrap}}
.logo{{display:flex;align-items:center;gap:12px}}
.logo-icon{{width:36px;height:36px;background:var(--naranja);border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:18px;font-weight:700}}
.logo-text{{font-size:16px;font-weight:600}}.logo-sub{{font-size:11px;opacity:.65}}
.hright{{display:flex;align-items:center;gap:10px}}
.corte-badge{{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.2);border-radius:20px;padding:5px 14px;font-size:12px;font-family:'DM Mono',monospace}}
.upd-badge{{font-size:11px;opacity:.55}}.user-badge{{font-size:12px;opacity:.8}}
.logout-btn{{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.2);color:#fff;border-radius:20px;padding:4px 12px;font-size:11px;cursor:pointer;font-family:'DM Sans',sans-serif}}
.logout-btn:hover{{background:rgba(255,255,255,.2)}}
main{{max-width:1200px;margin:0 auto;padding:2rem}}
.sec-title{{font-size:11px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;color:var(--muted);margin:2.5rem 0 1rem}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px}}
.kpi{{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:1.25rem;position:relative;overflow:hidden;transition:transform .15s,box-shadow .15s}}
.kpi:hover{{transform:translateY(-2px);box-shadow:0 8px 24px rgba(0,0,0,.08)}}
.kpi-accent{{position:absolute;top:0;left:0;right:0;height:3px}}
.kpi-label{{font-size:12px;color:var(--muted);margin-bottom:6px}}
.kpi-value{{font-size:26px;font-weight:600;letter-spacing:-1px;line-height:1}}
.kpi-delta{{display:inline-flex;align-items:center;gap:3px;font-size:12px;font-weight:500;font-family:'DM Mono',monospace;margin-top:8px;padding:2px 8px;border-radius:20px}}
.d-pos{{background:var(--pos-bg);color:var(--pos)}}.d-neg{{background:var(--neg-bg);color:var(--neg)}}.d-neu{{background:#f3f4f6;color:var(--muted)}}
.card{{background:var(--card);border:1px solid var(--border);border-radius:12px;overflow:hidden}}
.card-header{{padding:1rem 1.25rem;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid var(--border)}}
.card-header h3{{font-size:14px;font-weight:600}}.card-sub{{font-size:12px;color:var(--muted);font-family:'DM Mono',monospace}}
.tbl-wrap{{overflow-x:auto}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
thead th{{background:#f9fafb;padding:10px 14px;text-align:left;font-size:11px;font-weight:600;letter-spacing:.8px;text-transform:uppercase;color:var(--muted);border-bottom:1px solid var(--border)}}
thead th[colspan]{{text-align:center!important}}
thead th:not(:first-child):not([colspan]){{text-align:right}}
tbody tr{{border-bottom:1px solid #f3f4f6;transition:background .1s}}
tbody tr:hover{{background:#f9fafb}}tbody tr:last-child{{border-bottom:none}}
td{{padding:11px 14px}}td:not(:first-child){{text-align:right;font-family:'DM Mono',monospace}}
.suc-name{{font-weight:500;display:flex;align-items:center;gap:8px}}
.suc-dot{{width:8px;height:8px;border-radius:50%;flex-shrink:0}}
.num-bold{{font-weight:600}}.num-right{{text-align:right!important}}.num-center{{text-align:center!important;color:var(--muted);font-size:12px}}
.bold{{font-weight:600}}.muted{{color:var(--muted)}}
.bg-ops{{background:#f0f5fb}}.bg-kyc{{background:#f0f9f0}}
.bg-ops-l{{background:#f0f5fb}}.bg-kyc-l{{background:#f0f9f0}}.bg-red-l{{background:#fff5f5}}.bg-yel-l{{background:#fffbea}}
.total-row{{background:#f9fafb!important;font-weight:600}}
.pill{{display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:500;font-family:'DM Mono',monospace}}
.pill-pos{{background:var(--pos-bg);color:var(--pos)}}.pill-neg{{background:var(--neg-bg);color:var(--neg)}}.pill-neu{{background:#f3f4f6;color:var(--muted)}}
.doc-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:14px}}
.doc-card{{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:1.25rem}}
.doc-label{{font-size:12px;color:var(--muted);margin-bottom:4px}}.doc-pct{{font-size:26px;font-weight:600;margin-bottom:2px}}
.doc-count{{font-size:12px;color:var(--muted);font-family:'DM Mono',monospace}}
.progress{{width:100%;height:6px;background:#e5e7eb;border-radius:3px;margin-top:10px;overflow:hidden}}
.progress-bar{{height:100%;border-radius:3px}}
.donut-section{{display:grid;grid-template-columns:220px 1fr;gap:24px;align-items:center;padding:1.5rem}}
.donut-wrap{{position:relative;width:200px;height:200px}}
.donut-center{{position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);text-align:center;pointer-events:none}}
.donut-center-val{{font-size:20px;font-weight:600;letter-spacing:-1px}}
.donut-center-lbl{{font-size:11px;color:var(--muted)}}
.donut-legend{{display:flex;flex-direction:column;gap:6px}}
.donut-item{{display:flex;align-items:center;gap:8px;font-size:12px}}
.donut-dot{{width:10px;height:10px;border-radius:3px;flex-shrink:0}}
.donut-lbl{{flex:1}}.donut-val{{font-family:'DM Mono',monospace;font-size:11px;color:var(--muted);min-width:60px;text-align:right}}
.donut-pct{{font-family:'DM Mono',monospace;font-size:12px;font-weight:500;min-width:40px;text-align:right}}
.trans-row{{cursor:pointer}}.trans-row:hover{{background:#f0f5ff!important}}
.trans-row td:first-child::after{{content:" 🔍";font-size:10px;opacity:.4}}
.modal-overlay{{position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:1000;display:none;align-items:center;justify-content:center;padding:1rem}}
.modal-overlay.active{{display:flex}}
.modal-box{{background:var(--card);border-radius:16px;width:100%;max-width:700px;max-height:85vh;display:flex;flex-direction:column;box-shadow:0 20px 60px rgba(0,0,0,.3)}}
.modal-header{{padding:1.25rem 1.5rem;border-bottom:1px solid var(--border);display:flex;align-items:flex-start;justify-content:space-between;flex-shrink:0}}
.modal-title{{font-size:14px;font-weight:600;margin-bottom:3px}}.modal-sub{{font-size:12px;color:var(--muted)}}
.modal-close{{width:30px;height:30px;border-radius:8px;border:1px solid var(--border);background:none;cursor:pointer;font-size:15px;color:var(--muted)}}
.modal-body{{overflow-y:auto;flex:1}}.modal-body table{{font-size:13px}}
.modal-body thead th{{position:sticky;top:0;z-index:1}}
.modal-footer{{padding:.75rem 1.5rem;border-top:1px solid var(--border);font-size:12px;color:var(--muted);flex-shrink:0}}
.login-overlay{{position:fixed;inset:0;background:var(--azul);display:none;align-items:center;justify-content:center;z-index:9999}}
.login-overlay.active{{display:flex}}
.login-box{{background:#fff;border-radius:16px;padding:2.5rem 2rem;width:100%;max-width:340px;box-shadow:0 20px 60px rgba(0,0,0,.3)}}
.login-logo{{display:flex;align-items:center;gap:12px;margin-bottom:2rem;justify-content:center}}
.login-logo-icon{{width:44px;height:44px;background:var(--naranja);border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:22px;font-weight:700}}
.login-logo-text{{font-size:18px;font-weight:600;color:var(--azul)}}.login-logo-sub{{font-size:12px;color:var(--muted)}}
.login-field{{margin-bottom:1rem}}
.login-label{{font-size:12px;font-weight:500;color:var(--muted);margin-bottom:4px;display:block;text-transform:uppercase;letter-spacing:.8px}}
.login-input{{width:100%;padding:10px 14px;border:1.5px solid var(--border);border-radius:8px;font-size:14px;font-family:'DM Sans',sans-serif;outline:none;transition:border-color .15s;box-sizing:border-box}}
.login-input:focus{{border-color:var(--azul2)}}
.login-btn{{width:100%;padding:12px;background:var(--azul);color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;font-family:'DM Sans',sans-serif;margin-top:.5rem}}
.login-btn:hover{{background:#162d56}}
.login-error{{background:#fee2e2;color:var(--neg);border-radius:8px;padding:10px 14px;font-size:13px;margin-bottom:1rem;display:none}}
@media(max-width:640px){{main{{padding:1rem}}.donut-section{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<header>
  <div class="hinner">
    <div class="logo">
      <div class="logo-icon">C</div>
      <div><div class="logo-text">CECOM — Reportes PLD</div><div class="logo-sub">Prevención de Lavado de Dinero</div></div>
    </div>
    <div class="hright">
      <span class="corte-badge">{fa} vs {fc}</span>
      <span class="upd-badge">Actualizado: {gen}</span>
      <span class="user-badge" id="headerUser"></span>
      <button class="logout-btn" id="logoutBtn" onclick="doLogout()" style="display:none">Cerrar sesión</button>
    </div>
  </div>
</header>
<main>
  <div class="sec-title">Resumen global · {ns} sucursales</div>
  <div class="kpi-grid">
    <div class="kpi"><div class="kpi-accent" style="background:#1F3864"></div>
      <div class="kpi-label">Total clientes en base</div>
      <div class="kpi-value">{fmt(tc)}</div>
      <div class="kpi-delta d-pos">▲ +{fmt(k["delta_total"])} vs anterior</div>
    </div>
    <div class="kpi"><div class="kpi-accent" style="background:#E26B0A"></div>
      <div class="kpi-label">Clientes nuevos</div>
      <div class="kpi-value">{fmt(k["clientes_nuevos"])}</div>
      <div class="kpi-delta d-pos">▲ +{fmt(k["clientes_nuevos"])}</div>
    </div>
    <div class="kpi"><div class="kpi-accent" style="background:#7c3aed"></div>
      <div class="kpi-label">Cambios de estatus</div>
      <div class="kpi-value">{fmt(k["cambios_status"])}</div>
      <div class="kpi-delta d-neu">{k["cambios_status"]/tc*100:.1f}% de la base</div>
    </div>
    <div class="kpi"><div class="kpi-accent" style="background:#375623"></div>
      <div class="kpi-label">Verificados PLD</div>
      <div class="kpi-value">{fmt(k["verif_kyc_act"])}</div>
      <div class="kpi-delta d-pos">▲ +{fmt(k["verif_kyc_act"]-k["verif_kyc_ant"])} vs anterior</div>
    </div>
    <div class="kpi"><div class="kpi-accent" style="background:#dc2626"></div>
      <div class="kpi-label">Bloqueados PLD</div>
      <div class="kpi-value">{fmt(k["bloq_kyc_act"])}</div>
      <div class="kpi-delta d-{"neg" if k["bloq_kyc_act"]>k["bloq_kyc_ant"] else "pos"}">{"▲" if k["bloq_kyc_act"]>k["bloq_kyc_ant"] else "▼"} {fmt(abs(k["bloq_kyc_act"]-k["bloq_kyc_ant"]))} vs anterior</div>
    </div>
    <div class="kpi"><div class="kpi-accent" style="background:#0891b2"></div>
      <div class="kpi-label">Por validar PLD</div>
      <div class="kpi-value">{fmt(k["pend_kyc_act"])}</div>
      <div class="kpi-delta d-neg">▼ {fmt(abs(k["pend_kyc_act"]-k["pend_kyc_ant"]))} (mejora)</div>
    </div>
  </div>
  <div class="sec-title">Distribución de estatus · {fcl}</div>
  <div class="card">
    <div class="donut-section">
      <div class="donut-wrap" id="donutWrap">
        <div class="donut-center"><div class="donut-center-val">{fmt(tc)}</div><div class="donut-center-lbl">clientes totales</div></div>
      </div>
      <div class="donut-legend" id="donutLegend"></div>
    </div>
  </div>
  <div class="sec-title">Desglose por sucursal · {fcl}</div>
  <div class="card">
    <div class="card-header"><h3>Totales al {fcl}</h3><span class="card-sub">vs {fal}</span></div>
    <div class="tbl-wrap"><table>
      <thead>
        <tr>
          <th rowspan="2" style="vertical-align:middle">Sucursal</th>
          <th rowspan="2" style="vertical-align:middle;text-align:right">Total</th>
          <th colspan="4" style="text-align:center;background:#dce6f3;color:#1F3864">OPERACIONES</th>
          <th colspan="4" style="text-align:center;background:#d6e8d0;color:#375623">KYC-PLD</th>
          <th rowspan="2" style="vertical-align:middle;text-align:right">Reg. Express</th>
          <th rowspan="2" style="vertical-align:middle;text-align:right">Variación</th>
        </tr>
        <tr>
          <th style="background:#dce6f3;color:#1F3864;text-align:right">Bloqueado</th>
          <th style="background:#dce6f3;color:#1F3864;text-align:right">Info Act</th>
          <th style="background:#dce6f3;color:#1F3864;text-align:right">Reg Completo</th>
          <th style="background:#dce6f3;color:#1F3864;text-align:right">Verificado</th>
          <th style="background:#d6e8d0;color:#375623;text-align:right">Bloqueado</th>
          <th style="background:#d6e8d0;color:#375623;text-align:right">Info Act</th>
          <th style="background:#d6e8d0;color:#375623;text-align:right">Pend Valid</th>
          <th style="background:#d6e8d0;color:#375623;text-align:right">Verificado</th>
        </tr>
      </thead>
      <tbody>{kw["tabla_rows"]}</tbody>
    </table></div>
  </div>
  <div class="sec-title">Indicadores de documentación · {fcl}</div>
  <div class="doc-grid">
    <div class="doc-card">
      <div class="doc-label">Con Aviso de Privacidad</div>
      <div class="doc-pct">{pA}%</div>
      <div class="doc-count">{fmt(k["aviso_act"])} de {fmt(tc)}</div>
      <div class="progress"><div class="progress-bar" style="width:{pA}%;background:#2E75B6"></div></div>
    </div>
    <div class="doc-card">
      <div class="doc-label">Sin Aviso de Privacidad</div>
      <div class="doc-pct" style="color:var(--neg)">{pSA}%</div>
      <div class="doc-count">{fmt(tc-k["aviso_act"])} de {fmt(tc)}</div>
      <div class="progress"><div class="progress-bar" style="width:{pSA}%;background:#f97316"></div></div>
    </div>
    <div class="doc-card">
      <div class="doc-label">Con Huella Digital</div>
      <div class="doc-pct">{pH}%</div>
      <div class="doc-count">{fmt(k["huella_act"])} de {fmt(tc)}</div>
      <div class="progress"><div class="progress-bar" style="width:{pH}%;background:#375623"></div></div>
    </div>
    <div class="doc-card">
      <div class="doc-label">Sin Huella Digital</div>
      <div class="doc-pct" style="color:var(--neg)">{pSH}%</div>
      <div class="doc-count">{fmt(tc-k["huella_act"])} de {fmt(tc)}</div>
      <div class="progress"><div class="progress-bar" style="width:{pSH}%;background:#7c3aed"></div></div>
    </div>
    <div class="doc-card">
      <div class="doc-label">ID Vencida (≤ 31/12/2025)</div>
      <div class="doc-pct" style="color:var(--neg)">{pV}%</div>
      <div class="doc-count">{fmt(k["id_venc_act"])} de {fmt(tc)}</div>
      <div class="progress"><div class="progress-bar" style="width:{pV}%;background:#dc2626"></div></div>
    </div>
  </div>
  <div class="sec-title">Cambio de estatus · {fa} vs {fc}</div>
  <div class="card">
    <div class="card-header"><h3>Todos los movimientos</h3><span class="card-sub">{fmt(kw["total_c"])} cambios · ordenado mayor a menor</span></div>
    <div class="tbl-wrap"><table>
      <thead><tr>
        <th style="width:40px">#</th>
        <th>Status Anterior ({fa})</th>
        <th>Status Nuevo ({fc})</th>
        <th style="text-align:right">Cantidad</th>
        <th style="text-align:right">% del Total</th>
      </tr></thead>
      <tbody>{kw["trans_rows"]}</tbody>
    </table></div>
  </div>
  <div class="sec-title">Reporte de operaciones · {kw["fecha_ops_ini"]} al {kw["fecha_ops_fin"]}</div>
  <div class="card" style="padding:1.5rem;margin-bottom:14px">
    <canvas id="opsChart" style="max-height:300px"></canvas>
  </div>
  <div class="card" style="margin-bottom:14px">
    <div class="card-header"><h3>Clientes nuevos en el período — Operaciones</h3><span class="card-sub">{kw["fecha_ops_ini"]} – {kw["fecha_ops_fin"]}</span></div>
    <div class="tbl-wrap"><table>
      <thead><tr>
        <th>Sucursal</th><th style="text-align:right">Total</th>
        <th style="text-align:right;background:#f0f5fb">Reg. Completo</th>
        <th style="text-align:right;background:#f0f5fb">Verif. Ops</th>
        <th style="text-align:right;background:#fff5f5">Bloq. Ops</th>
        <th style="text-align:right;background:#f0f5fb">Info Act. Ops</th>
        <th style="text-align:right;background:#fffbea">Reg. Express</th>
      </tr></thead>
      <tbody>{kw["ops_rows"]}</tbody>
    </table></div>
  </div>
  <div class="card">
    <div class="card-header"><h3>Clientes nuevos en el período — KYC-PLD</h3><span class="card-sub">{kw["fecha_ops_ini"]} – {kw["fecha_ops_fin"]}</span></div>
    <div class="tbl-wrap"><table>
      <thead><tr>
        <th>Sucursal</th><th style="text-align:right">Total</th>
        <th style="text-align:right;background:#fff5f5">Bloq. KYC</th>
        <th style="text-align:right;background:#f0f9f0">Info Act. KYC</th>
        <th style="text-align:right;background:#f0f5fb">Pend. Valid. KYC</th>
        <th style="text-align:right;background:#f0f9f0">Verif. KYC</th>
      </tr></thead>
      <tbody>{kw["kyc_rows"]}</tbody>
    </table></div>
  </div>
</main>
<div class="modal-overlay" id="modalOverlay" onclick="if(event.target===this)cerrarModal()">
  <div class="modal-box">
    <div class="modal-header">
      <div><div class="modal-title" id="modalTitle"></div><div class="modal-sub" id="modalSub"></div></div>
      <button class="modal-close" onclick="cerrarModal()">✕</button>
    </div>
    <div class="modal-body">
      <table><thead><tr><th>Sucursal</th><th>Player ID</th></tr></thead>
      <tbody id="modalBody"></tbody></table>
    </div>
    <div class="modal-footer" id="modalFooter"></div>
  </div>
</div>
<div class="login-overlay active" id="loginOverlay">
  <div class="login-box">
    <div class="login-logo">
      <div class="login-logo-icon">C</div>
      <div><div class="login-logo-text">CECOM — PLD</div><div class="login-logo-sub">Portal de Reportes</div></div>
    </div>
    <div class="login-error" id="loginError">Usuario o contraseña incorrectos</div>
    <div class="login-field">
      <label class="login-label">Usuario</label>
      <input class="login-input" id="loginUser" type="text" placeholder="tu usuario" onkeydown="if(event.key==='Enter')doLogin()">
    </div>
    <div class="login-field">
      <label class="login-label">Contraseña</label>
      <input class="login-input" id="loginPass" type="password" placeholder="••••••••" onkeydown="if(event.key==='Enter')doLogin()">
    </div>
    <button class="login-btn" onclick="doLogin()">Iniciar sesión</button>
  </div>
</div>
<script>
const DONUT_DATA = {kw["donut_data"]};
const OPS_CHART  = {kw["ops_chart"]};
const DETALLE    = {kw["detalle_json"]};
const USUARIOS   = {kw["usuarios_js"]};
async function sha256(m){{const b=await crypto.subtle.digest('SHA-256',new TextEncoder().encode(m));return Array.from(new Uint8Array(b)).map(x=>x.toString(16).padStart(2,'0')).join('');}}
async function doLogin(){{
  const u=document.getElementById('loginUser').value.trim().toLowerCase();
  const p=document.getElementById('loginPass').value;
  const err=document.getElementById('loginError');
  if(!u||!p){{err.style.display='block';return;}}
  const h=await sha256(p);
  const found=USUARIOS.find(x=>x.u===u&&x.h===h);
  if(found){{
    sessionStorage.setItem('pld_user',JSON.stringify({{u:found.u,n:found.n}}));
    err.style.display='none';
    document.getElementById('loginOverlay').classList.remove('active');
    document.getElementById('headerUser').textContent=found.n;
    document.getElementById('logoutBtn').style.display='inline-block';
  }}else{{err.style.display='block';document.getElementById('loginPass').value='';document.getElementById('loginPass').focus();}}
}}
function doLogout(){{
  sessionStorage.removeItem('pld_user');
  document.getElementById('loginOverlay').classList.add('active');
  document.getElementById('headerUser').textContent='';
  document.getElementById('logoutBtn').style.display='none';
  document.getElementById('loginUser').value='';document.getElementById('loginPass').value='';
}}
window.addEventListener('DOMContentLoaded',()=>{{
  const s=sessionStorage.getItem('pld_user');
  if(s){{const u=JSON.parse(s);document.getElementById('loginOverlay').classList.remove('active');document.getElementById('headerUser').textContent=u.n;document.getElementById('logoutBtn').style.display='inline-block';}}
  initCharts();
}});
function initCharts(){{
  const total=DONUT_DATA.reduce((a,d)=>a+d.val,0);
  const circ=2*Math.PI*90;let offset=0;
  const paths=DONUT_DATA.map(d=>{{const dash=d.val/total*circ;const el=`<circle cx="110" cy="110" r="90" fill="none" stroke="${{d.color}}" stroke-width="28" stroke-dasharray="${{dash}} ${{circ-dash}}" stroke-dashoffset="${{-offset}}"/>`;offset+=dash;return el;}}).join('');
  const wrap=document.getElementById('donutWrap');
  wrap.innerHTML=`<svg viewBox="0 0 220 220" width="200" height="200" style="transform:rotate(-90deg)">${{paths}}</svg><div class="donut-center"><div class="donut-center-val">{fmt(tc)}</div><div class="donut-center-lbl">clientes totales</div></div>`;
  document.getElementById('donutLegend').innerHTML=DONUT_DATA.map(d=>{{const p=(d.val/total*100).toFixed(1);return `<div class="donut-item"><div class="donut-dot" style="background:${{d.color}}"></div><span class="donut-lbl">${{d.lbl}}</span><span class="donut-val">${{d.val.toLocaleString('es-MX')}}</span><span class="donut-pct" style="color:${{d.color}}">${{p}}%</span></div>`;}}).join('');
  if(typeof Chart!=='undefined'){{
    new Chart(document.getElementById('opsChart'),{{
      type:'bar',
      data:{{labels:OPS_CHART.labels,datasets:[
        {{label:'Reg. Completo',data:OPS_CHART.reg_completo,backgroundColor:'#1F3864'}},
        {{label:'Verif. Ops',data:OPS_CHART.verif_ops,backgroundColor:'#2E75B6'}},
        {{label:'Bloq. Ops',data:OPS_CHART.bloq_ops,backgroundColor:'#C00000'}},
        {{label:'Info Act.',data:OPS_CHART.info_act_ops,backgroundColor:'#70AD47'}},
        {{label:'Reg. Express',data:OPS_CHART.re,backgroundColor:'#FFC000'}},
      ]}},
      options:{{indexAxis:'y',responsive:true,
        plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{footer:i=>`Total: ${{OPS_CHART.totales[i[0].dataIndex].toLocaleString('es-MX')}}`,label:i=>` ${{i.dataset.label}}: ${{i.raw.toLocaleString('es-MX')}}`}}}}}},
        scales:{{x:{{stacked:true,ticks:{{callback:v=>v>=1000?(v/1000).toFixed(1)+'k':v,color:'#888780',font:{{size:11}}}}}},y:{{stacked:true,grid:{{display:false}},ticks:{{color:'#444441',font:{{size:12}}}}}}}}
      }}
    }});
  }}
}}
function abrirModal(tr){{
  const ant = tr.querySelector('[data-ant]')?.dataset.ant || '';
  const act = tr.querySelector('[data-act]')?.dataset.act || '';
  const f=DETALLE.filter(c=>c.ant===ant&&c.act===act);
  if(!f.length){{alert('No hay detalle disponible.');return;}}
  document.getElementById('modalTitle').textContent=ant+' → '+act;
  document.getElementById('modalSub').textContent=f.length.toLocaleString('es-MX')+' cliente(s)';
  document.getElementById('modalFooter').textContent=f.length.toLocaleString('es-MX')+' registros';
  document.getElementById('modalBody').innerHTML=f.map(c=>`<tr><td>${{c.sucursal}}</td><td style="font-family:'DM Mono',monospace;font-size:12px">${{c.id}}</td></tr>`).join('');
  document.getElementById('modalOverlay').classList.add('active');
  document.body.style.overflow='hidden';
}}
function cerrarModal(){{document.getElementById('modalOverlay').classList.remove('active');document.body.style.overflow='';}}
document.addEventListener('keydown',e=>{{if(e.key==='Escape')cerrarModal();}});
</script>
</body>
</html>"""
    return html


def generar_html(df_ant_all, df_act_all, datos_por_sucursal,
                 todos_cambios, todos_nuevos):
    """Copia el index.html plantilla desde la carpeta del script a COMPARATIVAS.

    El index.html NO se genera desde cero — se usa la plantilla maestra que
    carga los datos desde datos.json / cortes.json en tiempo de ejecución.
    Esto garantiza que el panel ⭐ Resumen, el selector de cortes y todas las
    mejoras visuales siempre estén presentes.
    """
    import shutil as _shutil

    # Buscar la plantilla en la carpeta del script (PORTAL WEB PLD\)
    script_dir = Path(__file__).resolve().parent
    plantilla  = script_dir / "index.html"

    ruta_html = CARPETA_SALIDA / "index.html"

    if plantilla.exists():
        _shutil.copy2(plantilla, ruta_html)
        print(f"  ✅ index.html copiado desde plantilla → {ruta_html}")
    else:
        # Si no existe la plantilla junto al script, no sobreescribir el existente
        if ruta_html.exists():
            print(f"  ℹ️  index.html no modificado (plantilla no encontrada en {script_dir})")
        else:
            print(f"  ⚠️  No se encontró plantilla en {script_dir}")
            print(f"      Copia manualmente el index.html a esa carpeta.")
    return  # Fin de generar_html



# ══════════════════════════════════════════════════════════════
# NOTIFICACIÓN POR WHATSAPP — CALLMEBOT
# ══════════════════════════════════════════════════════════════

def enviar_whatsapp(kpis_ant, kpis_act, sucursales):
    """Envía mensaje de WhatsApp vía CallMeBot al terminar el corte."""
    import urllib.request, urllib.parse, getpass

    apikey = WHATSAPP_APIKEY or getpass.getpass(
        "\n🔑 Ingresa tu API key de CallMeBot: ").strip()
    if not apikey:
        print("  ❌ No se ingresó API key.")
        return

    # ── Construir mensaje ────────────────────────────────────
    def fmt(n): return f"{int(n):,}"
    def dif(a, b, mejor_baja=False):
        d = int(a) - int(b)
        if d == 0: return "sin cambio"
        signo = "+" if d > 0 else ""
        if mejor_baja:
            emoji = "⚠️" if d > 0 else "✅"
        else:
            emoji = "✅" if d > 0 else "⚠️"
        return f"{emoji} {signo}{fmt(d)}"

    tc      = kpis_act["total"]
    v_act   = kpis_act["verif"];  v_ant = kpis_ant["verif"]
    b_act   = kpis_act["bloq"];   b_ant = kpis_ant["bloq"]
    p_act   = kpis_act["pend"];   p_ant = kpis_ant["pend"]
    nuevos  = kpis_act["nuevos"]
    cambios = kpis_act["cambios"]

    # Top 3 sucursales con más bloqueados
    top3 = sorted(sucursales, key=lambda s: s["bloq_kyc"], reverse=True)[:3]
    top3_txt = "\n".join([f"  • {s['nombre']}: {fmt(s['bloq_kyc'])} bloq." for s in top3])

    mensaje = (
        f"📊 *CECOM — Reporte PLD*\n"
        f"Corte: *{FECHA_ANTERIOR_CORTA} vs {FECHA_ACTUAL_CORTA}*\n"
        f"Actualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
        f"─────────────────\n"
        f"👥 Total clientes: *{fmt(tc)}* ({dif(tc, kpis_ant['total'])})\n"
        f"✔️  Verificados PLD: *{fmt(v_act)}* ({dif(v_act, v_ant)})\n"
        f"🔴 Bloqueados PLD: *{fmt(b_act)}* ({dif(b_act, b_ant, True)})\n"
        f"🟡 Por validar PLD: *{fmt(p_act)}* ({dif(p_act, p_ant, True)})\n"
        f"🆕 Clientes nuevos: *{fmt(nuevos)}*\n"
        f"🔄 Cambios estatus: *{fmt(cambios)}*\n"
        f"─────────────────\n"
        f"🏢 Top bloqueados:\n{top3_txt}\n"
        f"─────────────────\n"
        f"🔗 {URL_PORTAL}"
    )

    # ── Enviar a cada número ─────────────────────────────────
    numeros = WHATSAPP_NUMEROS if WHATSAPP_NUMEROS and WHATSAPP_NUMEROS[0] != "+5281XXXXXXXX" else []
    if not numeros:
        num = input("  📱 Número WhatsApp (ej. +5281XXXXXXXX): ").strip()
        if num: numeros = [num]

    print("\n📱 Enviando WhatsApp...")
    for numero in numeros:
        try:
            params = urllib.parse.urlencode({
                "phone":   numero,
                "text":    mensaje,
                "apikey":  apikey,
            })
            url = f"https://api.callmebot.com/whatsapp.php?{params}"
            with urllib.request.urlopen(url, timeout=10) as resp:
                respuesta = resp.read().decode("utf-8")
                if "Message sent" in respuesta or "200" in respuesta:
                    print(f"  ✅ Enviado a {numero}")
                else:
                    print(f"  ⚠️  Respuesta inesperada para {numero}: {respuesta[:80]}")
        except Exception as e:
            print(f"  ❌ Error al enviar a {numero}: {e}")


# ══════════════════════════════════════════════════════════════
# NOTIFICACIÓN POR CORREO — OUTLOOK
# ══════════════════════════════════════════════════════════════

def enviar_correo(kpis_ant, kpis_act, sucursales, errores):
    """Envía correo de notificación vía Outlook SMTP al terminar el corte."""
    import smtplib, getpass
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    pwd = CORREO_PASSWORD or getpass.getpass(
        f"\n🔑 Contraseña Outlook de {CORREO_REMITENTE}: ")

    # ── KPIs del correo ───────────────────────────────────────
    tc       = kpis_act["total"]
    delta_t  = tc - kpis_ant["total"]
    v_act    = kpis_act["verif"];  v_ant = kpis_ant["verif"]
    b_act    = kpis_act["bloq"];   b_ant = kpis_ant["bloq"]
    p_act    = kpis_act["pend"];   p_ant = kpis_ant["pend"]
    nuevos   = kpis_act["nuevos"]
    cambios  = kpis_act["cambios"]

    def fmt(n):  return f"{int(n):,}"
    def dif(a, b, mejor_baja=False):
        d = a - b
        if d == 0: return "sin cambio"
        signo = "+" if d > 0 else ""
        color = ""
        if mejor_baja:
            color = "#dc2626" if d > 0 else "#16a34a"
        else:
            color = "#16a34a" if d > 0 else "#dc2626"
        return f'<span style="color:{color};font-weight:600">{signo}{fmt(d)}</span>'

    # ── Top 3 sucursales con más bloqueados ───────────────────
    top_bloq = sorted(sucursales, key=lambda s: s["bloq_kyc"], reverse=True)[:3]
    top_rows = "".join([
        f'<tr><td style="padding:4px 10px">{s["nombre"]}</td>'
        f'<td style="padding:4px 10px;text-align:right;font-family:monospace">{fmt(s["bloq_kyc"])}</td></tr>'
        for s in top_bloq
    ])

    # ── HTML del correo ───────────────────────────────────────
    url_portal = "https://misaelu72-design.github.io/reportes-pld/"
    html = f"""
<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;background:#f4f6fa;margin:0;padding:20px">
<div style="max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.1)">

  <!-- Header -->
  <div style="background:#1F3864;padding:20px 24px;display:flex;align-items:center;gap:12px">
    <div style="background:#E26B0A;width:36px;height:36px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:18px;font-weight:700;color:#fff;flex-shrink:0">C</div>
    <div>
      <div style="color:#fff;font-size:15px;font-weight:600">CECOM — Reportes PLD</div>
      <div style="color:rgba(255,255,255,.65);font-size:11px">Prevención de Lavado de Dinero</div>
    </div>
  </div>

  <!-- Cuerpo -->
  <div style="padding:24px">
    <p style="margin:0 0 6px;font-size:13px;color:#6b7280">Corte generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}</p>
    <h2 style="margin:0 0 20px;font-size:18px;color:#1F3864">
      Reporte {FECHA_ANTERIOR_CORTA} vs {FECHA_ACTUAL_CORTA} — {len(sucursales)} sucursales
    </h2>

    <!-- KPIs -->
    <table style="width:100%;border-collapse:collapse;margin-bottom:20px">
      <tr>
        <td style="padding:12px;background:#f0f5fb;border-radius:8px;text-align:center;width:33%">
          <div style="font-size:11px;color:#6b7280;margin-bottom:4px">Total clientes</div>
          <div style="font-size:22px;font-weight:700;color:#1F3864">{fmt(tc)}</div>
          <div style="font-size:11px;margin-top:4px">{dif(delta_t, 0)} vs anterior</div>
        </td>
        <td style="width:2%"></td>
        <td style="padding:12px;background:#f0fdf4;border-radius:8px;text-align:center;width:33%">
          <div style="font-size:11px;color:#6b7280;margin-bottom:4px">Verificados PLD</div>
          <div style="font-size:22px;font-weight:700;color:#375623">{fmt(v_act)}</div>
          <div style="font-size:11px;margin-top:4px">{dif(v_act, v_ant)} vs anterior</div>
        </td>
        <td style="width:2%"></td>
        <td style="padding:12px;background:#fff5f5;border-radius:8px;text-align:center;width:33%">
          <div style="font-size:11px;color:#6b7280;margin-bottom:4px">Bloqueados PLD</div>
          <div style="font-size:22px;font-weight:700;color:#dc2626">{fmt(b_act)}</div>
          <div style="font-size:11px;margin-top:4px">{dif(b_act, b_ant, mejor_baja=True)} vs anterior</div>
        </td>
      </tr>
      <tr><td colspan="5" style="height:10px"></td></tr>
      <tr>
        <td style="padding:12px;background:#fffbea;border-radius:8px;text-align:center">
          <div style="font-size:11px;color:#6b7280;margin-bottom:4px">Por validar PLD</div>
          <div style="font-size:22px;font-weight:700;color:#d97706">{fmt(p_act)}</div>
          <div style="font-size:11px;margin-top:4px">{dif(p_act, p_ant, mejor_baja=True)} vs anterior</div>
        </td>
        <td></td>
        <td style="padding:12px;background:#f9fafb;border-radius:8px;text-align:center">
          <div style="font-size:11px;color:#6b7280;margin-bottom:4px">Clientes nuevos</div>
          <div style="font-size:22px;font-weight:700;color:#1F3864">{fmt(nuevos)}</div>
        </td>
        <td></td>
        <td style="padding:12px;background:#f9fafb;border-radius:8px;text-align:center">
          <div style="font-size:11px;color:#6b7280;margin-bottom:4px">Cambios de estatus</div>
          <div style="font-size:22px;font-weight:700;color:#1F3864">{fmt(cambios)}</div>
        </td>
      </tr>
    </table>

    <!-- Top bloqueados -->
    <div style="margin-bottom:20px">
      <div style="font-size:11px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:#6b7280;margin-bottom:8px">Top sucursales — Bloqueados PLD</div>
      <table style="width:100%;border-collapse:collapse;font-size:13px">
        <thead>
          <tr style="background:#f9fafb">
            <th style="padding:6px 10px;text-align:left;color:#6b7280;font-weight:500;font-size:11px">Sucursal</th>
            <th style="padding:6px 10px;text-align:right;color:#6b7280;font-weight:500;font-size:11px">Bloqueados</th>
          </tr>
        </thead>
        <tbody>{top_rows}</tbody>
      </table>
    </div>

    <!-- CTA -->
    <div style="text-align:center;margin-bottom:8px">
      <a href="{url_portal}" style="background:#1F3864;color:#fff;text-decoration:none;padding:12px 28px;border-radius:8px;font-size:13px;font-weight:600;display:inline-block">
        Ver portal completo →
      </a>
    </div>
  </div>

  <!-- Footer -->
  <div style="background:#f9fafb;padding:12px 24px;text-align:center;font-size:11px;color:#9ca3af;border-top:1px solid #e5e7eb">
    CECOM · Reporte generado automáticamente · No responder a este correo
  </div>
</div>
</body>
</html>"""

    # ── Armar y enviar ────────────────────────────────────────
    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"[PLD] Corte {FECHA_ANTERIOR_CORTA} vs {FECHA_ACTUAL_CORTA} — {fmt(tc)} clientes"
    msg["From"]    = CORREO_REMITENTE
    msg["To"]      = ", ".join(CORREO_DESTINATARIOS)
    msg.attach(MIMEText(html, "html", "utf-8"))

    try:
        print("\n📧 Enviando correo de notificación...")
        with smtplib.SMTP("smtp.office365.com", 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(CORREO_REMITENTE, pwd)
            smtp.sendmail(CORREO_REMITENTE, CORREO_DESTINATARIOS, msg.as_string())
        print(f"  ✅ Correo enviado a: {', '.join(CORREO_DESTINATARIOS)}")
    except smtplib.SMTPAuthenticationError:
        print("  ❌ Error de autenticación — verifica usuario y contraseña.")
    except Exception as e:
        print(f"  ❌ Error al enviar correo: {e}")

def main():
    CARPETA_SALIDA.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*55}")
    print(f"  REPORTE PLD — TODAS LAS SUCURSALES")
    print(f"  {FECHA_ANTERIOR_LARGA}  vs  {FECHA_ACTUAL_LARGA}")
    print(f"{'='*55}\n")

    # 0. Crear carpetas si no existen
    CARPETA_ANTERIOR.mkdir(parents=True, exist_ok=True)
    CARPETA_ACTUAL.mkdir(parents=True, exist_ok=True)
    CARPETA_SALIDA.mkdir(parents=True, exist_ok=True)

    # 1. Rotar ACTUAL → ANTERIOR (si hay archivos en ACTUAL)
    archivos_actual = list(CARPETA_ACTUAL.glob("*.xlsx")) + list(CARPETA_ACTUAL.glob("*.xlsm"))
    archivos_actual = [f for f in archivos_actual if not f.name.startswith("~$")]

    if not archivos_actual:
        print("⚠️  La carpeta ACTUAL está vacía.")
        print(f"   Pon los archivos *Altas.xlsx de esta semana en:")
        print(f"   {CARPETA_ACTUAL}")
        input("\nPresiona Enter para cerrar...")
        return

    # Preguntar fechas
    print("\n──────────────────────────────────────────")
    print("  Configuración del corte")
    print("──────────────────────────────────────────")
    print(f"  Archivos en ACTUAL: {len(archivos_actual)}")
    for f in archivos_actual:
        print(f"    • {f.name}")
    print()

    import re as _re, sys as _sys

    def _inferir_fecha(texto_corta):
        meses = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
                 "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12,
                 "Abr":4,"Ene":1,"Feb":2,"Mar":3,"Ago":8,"Sep":9,"Oct":10,"Nov":11,"Dic":12}
        m = _re.match(r"(\d{1,2})-([A-Za-z]{2,3})", texto_corta)
        if m:
            dia = int(m.group(1))
            mes = meses.get(m.group(2).capitalize(), 5)
            return pd.Timestamp(f"2026-{mes:02d}-{dia:02d}")
        return None

    def _inyectar_fechas(fa_corta, fa_larga, fc_corta, fc_larga):
        fecha_ant = _inferir_fecha(fa_corta)
        fecha_act = _inferir_fecha(fc_corta)
        sam_i = (fecha_ant - pd.Timedelta(days=6)) if fecha_ant else SEMANA_ANT_INICIO
        sam_f = fecha_ant if fecha_ant else SEMANA_ANT_FIN
        snv_i = (fecha_ant + pd.Timedelta(days=1)) if fecha_ant else SEMANA_NVA_INICIO
        snv_f = fecha_act if fecha_act else SEMANA_NVA_FIN
        _m = _sys.modules[__name__]
        _m.FECHA_ANTERIOR_CORTA = fa_corta
        _m.FECHA_ACTUAL_CORTA   = fc_corta
        _m.FECHA_ANTERIOR_LARGA = fa_larga
        _m.FECHA_ACTUAL_LARGA   = fc_larga
        _m.SEMANA_ANT_INICIO    = sam_i
        _m.SEMANA_ANT_FIN       = sam_f
        _m.SEMANA_NVA_INICIO    = snv_i
        _m.SEMANA_NVA_FIN       = snv_f

    def _correr_comparativa(carpeta_ant, fa_corta, fa_larga, fc_corta, fc_larga,
                             df_act_all_cache=None):
        """Corre una comparativa completa y devuelve df_act_all."""
        _inyectar_fechas(fa_corta, fa_larga, fc_corta, fc_larga)
        print(f"\n{'='*55}")
        print(f"  COMPARATIVA: {fa_corta} vs {fc_corta}")
        print(f"{'='*55}")

        dfs_ant, dfs_act, datos_por_sucursal, errores = [], [], [], []

        for suc in SUCURSALES:
            ruta_ant = buscar_archivo(carpeta_ant,   suc)
            ruta_act = buscar_archivo(CARPETA_ACTUAL, suc)
            if ruta_ant is None:
                errores.append((suc, "No encontrado en carpeta anterior")); continue
            if ruta_act is None:
                errores.append((suc, "No encontrado en carpeta actual")); continue
            try:
                df_ant = leer_df(ruta_ant, suc)
                df_act = df_act_all_cache[df_act_all_cache["_SUCURSAL"]==suc].copy() \
                         if df_act_all_cache is not None else leer_df(ruta_act, suc)
                dfs_ant.append(df_ant)
                dfs_act.append(df_act)
                datos_por_sucursal.append((suc, df_ant, df_act))
                print(f"  ✅ {suc:10}  ant={len(df_ant):>7,}  act={len(df_act):>7,}  Δ={len(df_act)-len(df_ant):+,}")
            except Exception as e:
                print(f"  ❌ {suc:10}  Error: {e}")
                errores.append((suc, str(e)))

        if not dfs_act:
            print("\n❌ No se pudo leer ningún archivo.")
            return None, None

        df_ant_all = pd.concat(dfs_ant, ignore_index=True)
        df_act_all = pd.concat(dfs_act, ignore_index=True)

        comun = df_ant_all[["_SUCURSAL","_ID","_STATUS","_NOTAS","_NOMBRE"]].merge(
            df_act_all[["_SUCURSAL","_ID","_STATUS","_NOTAS","_NOMBRE"]],
            on=["_SUCURSAL","_ID"], suffixes=("_ANT","_ACT"))
        todos_cambios = comun[comun["_STATUS_ANT"] != comun["_STATUS_ACT"]].copy()
        todos_nuevos  = df_act_all[
            ~df_act_all.set_index(["_SUCURSAL","_ID"]).index.isin(
             df_ant_all.set_index(["_SUCURSAL","_ID"]).index)].copy()

        print(f"\n  Total anterior : {len(df_ant_all):,}")
        print(f"  Total actual   : {len(df_act_all):,}")
        print(f"  Cambios status : {len(todos_cambios):,}")
        print(f"  Clientes nuevos: {len(todos_nuevos):,}")

        # Generar Excel
        print("\nGenerando reporte Excel...")
        wb = Workbook(); wb.remove(wb.active)
        pasos = [
            ("H1  - Cambio de Estatus",      lambda: hacer_h1(wb, todos_cambios, todos_nuevos)),
            ("H2  - Notas Bloqueados",        lambda: hacer_h2(wb, df_act_all)),
            ("H3  - Datos por Sucursal",      lambda: hacer_h3(wb, df_act_all, datos_por_sucursal)),
            ("H4  - Resumen Estatus",         lambda: hacer_h4(wb, df_ant_all, df_act_all, datos_por_sucursal)),
            ("H5  - Resumen Global",          lambda: hacer_h5(wb, df_ant_all, df_act_all, todos_cambios, todos_nuevos)),
            ("H6  - Detalle Reg Express",     lambda: hacer_h6(wb, todos_cambios, df_act_all)),
            ("H7  - Detalle Clientes Nuevos", lambda: hacer_h7(wb, todos_nuevos)),
            ("H8  - Resumen Registros Semana",lambda: hacer_h8(wb, df_ant_all, df_act_all)),
            ("H9  - Detalle Registros Semana",lambda: hacer_h9(wb, df_ant_all, df_act_all)),
            ("H10 - Resumen YTD 2026",        lambda: hacer_h10(wb, df_ant_all, df_act_all)),
            ("H11 - Detalle YTD 2026",        lambda: hacer_h11(wb, df_act_all)),
        ]
        for nom, fn in pasos:
            print(f"  {nom}..."); fn()

        nombre_xlsx = f"Reporte_PLD_TODAS_SUCURSALES_{fa_corta}_vs_{fc_corta}.xlsx"
        ruta_xlsx   = CARPETA_SALIDA / nombre_xlsx
        wb.save(ruta_xlsx)
        kb = ruta_xlsx.stat().st_size // 1024
        print(f"\n✅  {nombre_xlsx}  ({kb:,} KB)")

        # Generar JSON y HTML (solo en la comparativa principal)
        print("\nGenerando datos.json...")
        generar_json(df_ant_all, df_act_all, datos_por_sucursal, todos_cambios, todos_nuevos)

        print("\nGenerando detalle de cambios...")
        generar_detalle_cambios(todos_cambios, todos_nuevos, df_act_all)

        if errores:
            ruta_err = CARPETA_SALIDA / f"ERRORES_{fa_corta}_vs_{fc_corta}.xlsx"
            pd.DataFrame(errores, columns=["SUCURSAL","ERROR"]).to_excel(ruta_err, index=False)
            print(f"\n⚠  {len(errores)} error(es) — ver: {ruta_err}")

        return df_act_all, todos_cambios

    # ── Pedir solo la fecha ACTUAL ────────────────────────────
    print("Solo necesitas ingresar la fecha del corte ACTUAL:")
    _fc_corta = input("  Fecha ACTUAL corta (ej. 30-May): ").strip()
    _fc_larga = input("  Fecha ACTUAL larga (ej. 30 de Mayo 2026): ").strip()

    # Detectar fecha anterior desde cortes.json
    # Buscar el corte más reciente que NO sea de 26-Abr y tomar su fecha ACTUAL
    # como la nueva fecha ANTERIOR (ej. "30-May vs 07-Jun" → anterior = "30-May")
    _ruta_cortes = CARPETA_SALIDA / "cortes.json"
    _fa_corta, _fa_larga = "Ant", "Corte Anterior"
    _meses_es = {"Jan":"Enero","Feb":"Febrero","Mar":"Marzo","Apr":"Abril",
                 "May":"Mayo","Jun":"Junio","Jul":"Julio","Aug":"Agosto",
                 "Sep":"Septiembre","Oct":"Octubre","Nov":"Noviembre","Dec":"Diciembre",
                 "Abr":"Abril","Ene":"Enero","Mar":"Marzo","Dic":"Diciembre"}
    if _ruta_cortes.exists():
        import json as _json
        _cortes_prev = _json.loads(_ruta_cortes.read_text(encoding="utf-8"))
        # Filtrar cortes que no sean de referencia fija (26-Abr vs 26-Abr o similar)
        # Buscar cortes que NO sean la referencia fija de Abril
        # y cuya fecha ANTERIOR no sea "26-Abr"
        _cortes_norm = [c for c in _cortes_prev
                        if " vs " in c["etiqueta"]
                        and not c["etiqueta"].startswith("26 de")
                        and not c["etiqueta"].startswith("26-Abr")
                        and c["etiqueta"].count(" vs ") == 1]
        if _cortes_norm:
            _etiq = _cortes_norm[0]["etiqueta"]  # el más reciente
            partes = _etiq.split(" vs ")
            if len(partes) == 2:
                # La fecha ANTERIOR del próximo corte = fecha ACTUAL del corte más reciente
                _fa_corta = partes[1].strip()
                m = _re.match(r"(\d{1,2})-([A-Za-z]{2,3})", _fa_corta)
                if m:
                    _fa_larga = f"{m.group(1)} de {_meses_es.get(m.group(2).capitalize(), m.group(2))} 2026"

    # Mostrar la fecha detectada y permitir corregirla
    print(f"\n  Fecha anterior detectada: {_fa_corta}")
    _confirm = input(f"  ¿Es correcta? Enter para confirmar o escribe otra (ej. 30-May): ").strip()
    if _confirm:
        _fa_corta = _confirm
        m = _re.match(r"(\d{1,2})-([A-Za-z]{2,3})", _fa_corta)
        if m:
            _fa_larga = f"{m.group(1)} de {_meses_es.get(m.group(2).capitalize(), m.group(2))} 2026"

    print(f"\n  Comparativa 1: {_fa_corta} vs {_fc_corta}  (corte anterior vs actual)")
    print(f"  Comparativa 2: 26-Abr vs {_fc_corta}  (referencia fija vs actual)")

    # ── COMPARATIVA 1: ANTERIOR vs ACTUAL ────────────────────
    df_act_all, _ = _correr_comparativa(
        CARPETA_ANTERIOR, _fa_corta, _fa_larga, _fc_corta, _fc_larga
    )

    # ── COMPARATIVA 2: 26-ABR vs ACTUAL ──────────────────────
    if CARPETA_ABRIL.exists():
        _correr_comparativa(
            CARPETA_ABRIL, "26-Abr", "26 de Abril 2026", _fc_corta, _fc_larga,
            df_act_all_cache=df_act_all  # reutiliza los datos actuales ya leídos
        )
    else:
        print(f"\n⚠  Carpeta 04-26 no encontrada en {CARPETA_ABRIL} — omitiendo comparativa con Abril.")

    # ── Base unificada (una sola vez con datos actuales) ─────
    _inyectar_fechas(_fa_corta, _fa_larga, _fc_corta, _fc_larga)
    print("\nGenerando base unificada...")
    generar_base_unificada(df_act_all)

    # ── index.html (una sola vez — usa el datos.json del corte principal) ──
    print("\nGenerando index.html...")
    generar_html(None, None, None, None, None)

    print(f"\n{'='*55}")
    print("  PROCESO TERMINADO")
    print(f"{'='*55}\n")


    # ── Notificación por correo ─────────────────────────────
    respuesta_correo = input("¿Enviar notificación por correo? (s/n): ").strip().lower()
    if respuesta_correo in ["s", "si", "sí", "y", "yes"]:
        import json as _json2
        _ruta_datos = CARPETA_SALIDA / "datos.json"
        _d = _json2.loads(_ruta_datos.read_text(encoding="utf-8")) if _ruta_datos.exists() else {}
        _kpis = _d.get("kpis", {})
        kpis_ant = {
            "total":   _kpis.get("total_ant", 0),
            "verif":   _kpis.get("verif_kyc_ant", 0),
            "bloq":    _kpis.get("bloq_kyc_ant", 0),
            "pend":    _kpis.get("pend_kyc_ant", 0),
            "nuevos":  _kpis.get("clientes_nuevos", 0),
            "cambios": _kpis.get("cambios_status", 0),
        }
        kpis_act = {
            "total":   _kpis.get("total_act", 0),
            "verif":   _kpis.get("verif_kyc_act", 0),
            "bloq":    _kpis.get("bloq_kyc_act", 0),
            "pend":    _kpis.get("pend_kyc_act", 0),
            "nuevos":  _kpis.get("clientes_nuevos", 0),
            "cambios": _kpis.get("cambios_status", 0),
        }
        suc_list = [
            {"nombre": s["nombre"], "bloq_kyc": s.get("bloq_kyc_act", 0)}
            for s in _d.get("sucursales", [])
        ]
        enviar_correo(kpis_ant, kpis_act, suc_list, errores)

    # ── Notificación WhatsApp ────────────────────────────────
    resp_wa = input("¿Enviar notificación por WhatsApp? (s/n): ").strip().lower()
    if resp_wa in ["s", "si", "sí", "y", "yes"]:
        import json as _json3
        _ruta_d = CARPETA_SALIDA / "datos.json"
        _d2 = _json3.loads(_ruta_d.read_text(encoding="utf-8")) if _ruta_d.exists() else {}
        _k2 = _d2.get("kpis", {})
        kpis_ant_wa = {
            "total":   _k2.get("total_ant", 0),
            "verif":   _k2.get("verif_kyc_ant", 0),
            "bloq":    _k2.get("bloq_kyc_ant", 0),
            "pend":    _k2.get("pend_kyc_ant", 0),
            "nuevos":  _k2.get("clientes_nuevos", 0),
            "cambios": _k2.get("cambios_status", 0),
        }
        kpis_act_wa = {
            "total":   _k2.get("total_act", 0),
            "verif":   _k2.get("verif_kyc_act", 0),
            "bloq":    _k2.get("bloq_kyc_act", 0),
            "pend":    _k2.get("pend_kyc_act", 0),
            "nuevos":  _k2.get("clientes_nuevos", 0),
            "cambios": _k2.get("cambios_status", 0),
        }
        suc_list_wa = [
            {"nombre": s["nombre"], "bloq_kyc": s.get("bloq_kyc_act", 0)}
            for s in _d2.get("sucursales", [])
        ]
        enviar_whatsapp(kpis_ant_wa, kpis_act_wa, suc_list_wa)

    # ── Subida automática a GitHub ──────────────────────────
    script_github = Path(__file__).resolve().parent / "subir_a_github.py"
    if script_github.exists():
        respuesta = input("¿Subir automáticamente a GitHub? (s/n): ").strip().lower()
        if respuesta in ["s", "si", "sí", "y", "yes"]:
            import subprocess, sys
            subprocess.run([sys.executable, str(script_github)], check=False)
    else:
        print("ℹ️  Coloca 'subir_a_github.py' en la misma carpeta para subida automática.")


if __name__ == "__main__":
    main()